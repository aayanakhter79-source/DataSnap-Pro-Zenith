"""
DataSnap Pro - AI-Powered Financial OS for Freelancers
Built by Zenith IN | Version 2.0.0
FIXES: Persistence · AI Bridge · Client-Wise · Excel · GSheets stub
"""

import streamlit as st
import pandas as pd
import io, hashlib, base64, json
from datetime import datetime, date
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── PAGE CONFIG — must be FIRST ─────────────────────────────────────────────
st.set_page_config(page_title="DataSnap Pro | Zenith IN", page_icon="⚡",
                   layout="wide", initial_sidebar_state="expanded")

# ─── FIX #1: SINGLE INIT — safe across every rerun ───────────────────────────
def init_state():
    for k, v in {
        "invoices":        [],
        "current_invoice": None,
        "ai_extracted":    None,
        "authenticated":   False,
        "username":        "",
        "role":            "Client",
        "display_name":    "User",
    }.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ─── CONSTANTS ────────────────────────────────────────────────────────────────
USD_TO_INR = 93.08
CGST_RATE  = 0.09
SGST_RATE  = 0.09
TDS_RATES  = {
    "194J – Professional / Technical (10%)":  0.10,
    "194J – Royalty / FTS (2%)":              0.02,
    "194C – Contractor Individual/HUF (1%)":  0.01,
    "194C – Contractor Company (2%)":         0.02,
    "No TDS":                                 0.00,
}

# ─── USERS ───────────────────────────────────────────────────────────────────
USERS = {
    "admin":   {"password": hashlib.sha256(b"zenith@2026").hexdigest(),  "role": "Admin",  "name": "Admin User"},
    "client1": {"password": hashlib.sha256(b"client1pass").hexdigest(), "role": "Client", "name": "Rahul Sharma"},
    "client2": {"password": hashlib.sha256(b"client2pass").hexdigest(), "role": "Client", "name": "Priya Mehta"},
}

# ─── FIX #6: GOOGLE SHEETS STUB ──────────────────────────────────────────────
# To enable: uncomment below + add gspread to requirements.txt
# + add [gcp_service_account] and SHEET_ID to Streamlit Secrets
#
# import gspread
# from google.oauth2.service_account import Credentials
#
# @st.cache_resource
# def get_gsheet():
#     creds = Credentials.from_service_account_info(
#         st.secrets["gcp_service_account"],
#         scopes=["https://www.googleapis.com/auth/spreadsheets"])
#     return gspread.authorize(creds).open_by_key(st.secrets["SHEET_ID"]).sheet1
#
# def _push_to_gsheet(inv):
#     get_gsheet().append_row([
#         inv["date"], inv.get("client_name",""), inv.get("description",""),
#         inv["mode"], inv.get("usd_amount",""), inv["inr_amount"],
#         inv["cgst"], inv["sgst"], inv["total_gst"],
#         inv.get("invoice_total", inv["inr_amount"]),
#         inv["tds_amount"], inv["net_receivable"]
#     ])

def _push_to_gsheet(inv):
    pass   # STUB — wire the real function above when ready


def save_invoice(inv: dict):
    """Append to master list + optional cloud sync."""
    st.session_state["invoices"].append(inv)
    _push_to_gsheet(inv)


# ─── GEMINI ──────────────────────────────────────────────────────────────────
@st.cache_resource
def get_gemini():
    try:
        genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
        models = [m.name for m in genai.list_models()
                  if "generateContent" in m.supported_generation_methods]
        name = "models/gemini-1.5-flash" if "models/gemini-1.5-flash" in models else (models[0] if models else None)
        return genai.GenerativeModel(name) if name else None
    except Exception:
        return None


# ─── CSS ─────────────────────────────────────────────────────────────────────
def inject_css():
    st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@300;400;500&display=swap');
:root{--black:#080c10;--surface:#0d1117;--card:#111820;--border:#1e2d3d;
      --blue:#00a3ff;--blue2:#0066cc;--white:#f0f6fc;--muted:#8b98a5;
      --green:#00e676;--red:#ff5252;--gold:#ffd700;}
html,body,[data-testid="stAppViewContainer"]{background-color:var(--black)!important;color:var(--white)!important;font-family:'DM Mono',monospace!important;}
[data-testid="stSidebar"]{background:var(--surface)!important;border-right:1px solid var(--border)!important;}
[data-testid="stSidebar"] *{color:var(--white)!important;}
h1,h2,h3{font-family:'Syne',sans-serif!important;font-weight:800!important;}
.metric-card{background:var(--card);border:1px solid var(--border);border-top:3px solid var(--blue);border-radius:8px;padding:1.2rem 1.4rem;margin-bottom:.8rem;}
.metric-label{font-size:.72rem;letter-spacing:.12em;color:var(--muted);text-transform:uppercase;}
.metric-value{font-size:1.9rem;font-weight:700;font-family:'Syne',sans-serif;color:var(--white);line-height:1.2;}
.metric-sub{font-size:.75rem;color:var(--blue);margin-top:.2rem;}
.badge-export{background:#00332a;color:var(--green);padding:2px 8px;border-radius:4px;font-size:.7rem;font-weight:600;border:1px solid var(--green);}
.badge-inr{background:#2a1a00;color:var(--gold);padding:2px 8px;border-radius:4px;font-size:.7rem;font-weight:600;border:1px solid var(--gold);}
.badge-admin{background:#001833;color:var(--blue);padding:2px 8px;border-radius:4px;font-size:.7rem;font-weight:600;border:1px solid var(--blue);}
.section-header{border-left:3px solid var(--blue);padding-left:.8rem;margin:1.5rem 0 1rem;font-family:'Syne',sans-serif;font-size:1.1rem;font-weight:700;color:var(--white);}
.stButton>button{background:linear-gradient(135deg,var(--blue2),var(--blue))!important;color:#fff!important;border:none!important;border-radius:6px!important;font-family:'Syne',sans-serif!important;font-weight:700!important;letter-spacing:.05em!important;padding:.55rem 1.4rem!important;transition:opacity .2s!important;}
.stButton>button:hover{opacity:.85!important;}
[data-testid="stTextInput"] input,[data-testid="stNumberInput"] input,.stTextArea textarea{background:var(--card)!important;color:var(--white)!important;border:1px solid var(--border)!important;border-radius:6px!important;font-family:'DM Mono',monospace!important;}
[data-testid="stDataFrame"]{background:var(--card)!important;}
.zenith-logo{font-family:'Syne',sans-serif;font-size:1.4rem;font-weight:800;letter-spacing:.08em;color:var(--blue);}
.zenith-sub{font-size:.68rem;color:var(--muted);letter-spacing:.15em;text-transform:uppercase;}
.whatsapp-preview{background:#0a1a0f;border:1px solid #1a4a1a;border-radius:8px;padding:1rem 1.2rem;font-size:.82rem;color:#cce8cc;white-space:pre-wrap;font-family:'DM Mono',monospace;}
.ai-response{background:var(--card);border:1px solid var(--border);border-left:3px solid var(--blue);border-radius:8px;padding:1rem 1.2rem;font-size:.85rem;line-height:1.6;color:var(--white);}
.bridge-card{background:#001a2e;border:1px solid #004080;border-radius:8px;padding:1rem 1.2rem;margin-top:.8rem;}
.stTabs [data-baseweb="tab"]{background:transparent!important;color:var(--muted)!important;font-family:'Syne',sans-serif!important;font-weight:600!important;border-radius:0!important;border-bottom:2px solid transparent!important;}
.stTabs [aria-selected="true"]{color:var(--blue)!important;border-bottom:2px solid var(--blue)!important;background:transparent!important;}
div[data-testid="stAlert"]{border-radius:6px!important;}
footer{display:none!important;}
</style>""", unsafe_allow_html=True)


# ─── HELPERS ─────────────────────────────────────────────────────────────────
fmt_inr = lambda v: f"₹{float(v):,.2f}"
fmt_usd = lambda v: f"${float(v):,.2f}"
hash_pw = lambda pw: hashlib.sha256(pw.encode()).hexdigest()


# ─── CORE CALCULATOR ─────────────────────────────────────────────────────────
def calculate_invoice(mode, amount, tds_label, description, client_name):
    tds_rate = TDS_RATES[tds_label]
    if mode == "USD (Export)":
        inr = amount * USD_TO_INR
        inv = dict(mode="USD – Export of Service", usd_amount=amount, inr_amount=inr,
                   gst_type="LUT / Export – 0% GST", cgst=0.0, sgst=0.0, total_gst=0.0,
                   taxable_value=inr, tds_rate=tds_rate, tds_amount=inr*tds_rate,
                   net_receivable=inr*(1-tds_rate), exchange_rate=USD_TO_INR,
                   invoice_total=inr)
    else:
        cgst = amount * CGST_RATE; sgst = amount * SGST_RATE
        total_gst = cgst + sgst; inv_total = amount + total_gst
        inv = dict(mode="INR – Domestic", usd_amount=None, inr_amount=amount,
                   gst_type="18% GST (CGST 9% + SGST 9%)", cgst=cgst, sgst=sgst,
                   total_gst=total_gst, taxable_value=amount, invoice_total=inv_total,
                   tds_rate=tds_rate, tds_amount=amount*tds_rate,
                   net_receivable=inv_total - amount*tds_rate, exchange_rate=None)
    inv.update(description=description,
               client_name=client_name.strip() or "Unknown",
               date=datetime.now().strftime("%d %b %Y"),
               saved_by=st.session_state.get("username",""))
    return inv


# ─── WHATSAPP ────────────────────────────────────────────────────────────────
def whatsapp_message(inv, client_name, phone):
    lines = [f"👋 Hello {client_name}!","",
             "📋 *Invoice Summary — DataSnap Pro*",
             f"🗓  Date: {inv['date']}",f"📝 Desc: {inv.get('description','—')}","",
             f"💰 Mode: {inv['mode']}"]
    if inv.get("usd_amount"):
        lines.append(f"   USD: {fmt_usd(inv['usd_amount'])} → {fmt_inr(inv['inr_amount'])}")
    else:
        lines += [f"   Taxable: {fmt_inr(inv['inr_amount'])}",
                  f"   GST (18%): {fmt_inr(inv['total_gst'])}",
                  f"   Invoice Total: {fmt_inr(inv.get('invoice_total',inv['inr_amount']))}"]
    if inv["tds_amount"] > 0:
        lines.append(f"   TDS: {fmt_inr(inv['tds_amount'])} ({inv['tds_rate']*100:.0f}%)")
    lines += ["",f"✅ *Net Receivable: {fmt_inr(inv['net_receivable'])}*","",
              "Powered by ⚡ Zenith IN / DataSnap Pro"]
    return "\n".join(lines)

def send_whatsapp_placeholder(phone, message):
    # Twilio: from twilio.rest import Client; Client(sid,tok).messages.create(from_='whatsapp:+14155238886',to=f'whatsapp:{phone}',body=message)
    # WATI:   requests.post(url,headers={"Authorization":f"Bearer {tok}"},json={"messageText":message})
    return {"status":"mock_sent","to":phone,"chars":len(message)}


# ─── FIX #5: EXCEL EXPORT ────────────────────────────────────────────────────
def build_excel(invoices: list, owner: str) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "CA Audit Report"
    D="0D1117"; H="0A2540"; B="00A3FF"; W="F0F6FC"
    M="8B98A5"; G="00E676"; GO="FFD700"; R="FF5252"

    def cs(sheet, r, c, val, bold=False, fg=W, bg=D, al="left", nf=None, sz=10):
        cell = sheet.cell(row=r, column=c, value=val)
        cell.font = Font(name="Calibri", bold=bold, color=fg, size=sz)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=al, vertical="center", wrap_text=True)
        cell.border = Border(**{s: Side(style="thin", color="1E2D3D") for s in ["left","right","top","bottom"]})
        if nf: cell.number_format = nf
        return cell

    # Title
    ws.merge_cells("A1:M1")
    cs(ws,1,1,"⚡  DataSnap Pro  |  CA Audit Report  |  Zenith IN",bold=True,fg=B,bg="080C10",al="center",sz=14)
    ws.row_dimensions[1].height=30
    ws.merge_cells("A2:M2")
    cs(ws,2,1,f"For: {owner}   |   {datetime.now().strftime('%d %b %Y %H:%M')}   |   Rate: ₹{USD_TO_INR}/USD",
       fg=M,bg="080C10",al="center",sz=9)
    ws.row_dimensions[2].height=18

    hdrs=["#","Date","Client","Description","Mode","USD","INR (Taxable)","CGST 9%","SGST 9%","GST Total","Invoice Total","TDS","Net Receivable"]
    wds= [5,   14,    18,      30,            20,    14,   16,            12,       12,       12,         14,            12,   16]
    for i,(h,w) in enumerate(zip(hdrs,wds),1):
        cs(ws,3,i,h,bold=True,fg=B,bg=H,al="center",sz=10)
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[3].height=22

    tot=dict(usd=0,inr=0,cgst=0,sgst=0,gst=0,ivt=0,tds=0,net=0)
    for idx,inv in enumerate(invoices,1):
        r=idx+3; bg=D if idx%2==0 else "0F161F"
        is_usd=inv.get("usd_amount") is not None
        ivt=inv.get("invoice_total",inv["inr_amount"])
        vals=[idx,inv["date"],inv.get("client_name","—"),inv.get("description","—"),inv["mode"],
              inv["usd_amount"] if is_usd else "—",inv["inr_amount"],inv["cgst"],inv["sgst"],
              inv["total_gst"],ivt,inv["tds_amount"],inv["net_receivable"]]
        fmts=[None,None,None,None,None,"[$$-en-US]#,##0.00","₹#,##0.00","₹#,##0.00","₹#,##0.00","₹#,##0.00","₹#,##0.00","₹#,##0.00","₹#,##0.00"]
        fgs=[M,M,W,W,(GO if is_usd else G),W,W,W,W,W,W,R,G]
        for ci,(v,nf,fg) in enumerate(zip(vals,fmts,fgs),1):
            cs(ws,r,ci,v,fg=fg,bg=bg,nf=nf,al="right" if ci>5 else "left")
        tot["usd"]+=inv.get("usd_amount") or 0; tot["inr"]+=inv["inr_amount"]
        tot["cgst"]+=inv["cgst"]; tot["sgst"]+=inv["sgst"]; tot["gst"]+=inv["total_gst"]
        tot["ivt"]+=ivt; tot["tds"]+=inv["tds_amount"]; tot["net"]+=inv["net_receivable"]

    tr=len(invoices)+4; ws.row_dimensions[tr].height=22
    tv=["","TOTALS","","","",tot["usd"],tot["inr"],tot["cgst"],tot["sgst"],tot["gst"],tot["ivt"],tot["tds"],tot["net"]]
    tf=[None,None,None,None,None,"[$$-en-US]#,##0.00","₹#,##0.00","₹#,##0.00","₹#,##0.00","₹#,##0.00","₹#,##0.00","₹#,##0.00","₹#,##0.00"]
    for ci,(v,nf) in enumerate(zip(tv,tf),1):
        cs(ws,tr,ci,v,bold=True,fg=B,bg=H,nf=nf,al="right" if ci>5 else "left",sz=11)

    # FIX #3: Client-wise sheet
    ws3=wb.create_sheet("Client Summary"); ws3.sheet_view.showGridLines=False
    clients={}
    for inv in invoices:
        cn=inv.get("client_name","Unknown")
        if cn not in clients: clients[cn]={"rev":0,"gst":0,"tds":0,"net":0,"n":0}
        clients[cn]["rev"]+=inv["inr_amount"]; clients[cn]["gst"]+=inv["total_gst"]
        clients[cn]["tds"]+=inv["tds_amount"]; clients[cn]["net"]+=inv["net_receivable"]
        clients[cn]["n"]+=1
    for i,(h,w) in enumerate(zip(["Client","Invoices","Revenue","GST","TDS","Net"],[25,10,18,14,14,18]),1):
        cs(ws3,1,i,h,bold=True,fg=B,bg=H,al="center")
        ws3.column_dimensions[get_column_letter(i)].width=w
    for ri,(cn,d) in enumerate(clients.items(),2):
        bg2=D if ri%2==0 else "0F161F"
        for ci,(v,nf) in enumerate(zip([cn,d["n"],d["rev"],d["gst"],d["tds"],d["net"]],
                                       [None,None,"₹#,##0.00","₹#,##0.00","₹#,##0.00","₹#,##0.00"]),1):
            cs(ws3,ri,ci,v,fg=W,bg=bg2,nf=nf,al="right" if ci>1 else "left")
        ws3.row_dimensions[ri].height=20

    # Summary sheet
    ws2=wb.create_sheet("Summary"); ws2.sheet_view.showGridLines=False
    sd=[("METRIC","AMOUNT","NOTES"),("Total Invoices",len(invoices),"Count"),
        ("Total Revenue (INR)",tot["inr"],"Taxable"),("Total GST",tot["gst"],"CGST+SGST"),
        ("  ↳ CGST",tot["cgst"],"Central"),("  ↳ SGST",tot["sgst"],"State"),
        ("TDS Receivable",tot["tds"],"Claim ITR"),("Net Receivable",tot["net"],"After TDS"),
        ("USD Revenue",tot["usd"],"Export LUT")]
    for ri,row in enumerate(sd,1):
        ih=ri==1
        bg3=H if ih else (D if ri%2==0 else "0F161F")
        for ci,val in enumerate(row,1):
            fg3=B if ih else (G if ci==2 and ri>1 else W)
            nf3="₹#,##0.00" if ci==2 and ri not in[1,2,9] else ("[$$-en-US]#,##0.00" if ci==2 and ri==9 else None)
            cs(ws2,ri,ci,val,bold=ih,fg=fg3,bg=bg3,nf=nf3,al="right" if ci==2 else "left")
        ws2.row_dimensions[ri].height=20
    for ci,w in[(1,30),(2,20),(3,30)]: ws2.column_dimensions[get_column_letter(ci)].width=w

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


# ─── LOGIN ───────────────────────────────────────────────────────────────────
def login_block():
    st.markdown("<div style='text-align:center;padding:3rem 0 1.5rem'><div class='zenith-logo'>⚡ ZENITH IN</div><div class='zenith-sub'>DataSnap Pro · Financial OS v2.0</div></div>",
                unsafe_allow_html=True)
    col=st.columns([1,1.2,1])[1]
    with col:
        st.markdown("#### Sign In")
        u=st.text_input("Username",key="li_user")
        p=st.text_input("Password",type="password",key="li_pw")
        if st.button("LOGIN  →",use_container_width=True):
            usr=USERS.get(u)
            if usr and usr["password"]==hash_pw(p):
                st.session_state.update({"authenticated":True,"username":u,
                                          "role":usr["role"],"display_name":usr["name"]})
                st.rerun()
            else: st.error("Invalid credentials.")
        st.markdown("<div style='font-size:.72rem;color:#8b98a5;margin-top:1rem'>Demo → admin / zenith@2026<br>Demo → client1 / client1pass</div>",
                    unsafe_allow_html=True)


# ─── SIDEBAR ─────────────────────────────────────────────────────────────────
def render_sidebar():
    with st.sidebar:
        st.markdown("<div style='padding:1rem 0 .5rem'><div class='zenith-logo'>⚡ ZENITH IN</div><div class='zenith-sub'>DataSnap Pro v2.0</div></div>",
                    unsafe_allow_html=True)
        st.divider()
        name=st.session_state["display_name"]; role=st.session_state["role"]
        st.markdown(f"<div style='margin-bottom:1rem'><div style='font-size:.8rem;color:#8b98a5'>Logged in as</div><div style='font-size:1rem;font-weight:700'>{name}</div><span class='{'badge-admin' if role=='Admin' else 'badge-inr'}'>{role}</span></div>",
                    unsafe_allow_html=True)
        pages=["📊 Dashboard","🧾 New Invoice","📁 Invoice History","🤖 AI Assistant","📤 Export & WhatsApp"]
        if role=="Admin": pages.append("👥 Admin Panel")
        page=st.radio("Navigation",pages,label_visibility="collapsed")
        st.divider()
        n=len(st.session_state["invoices"])
        st.markdown(f"<div style='font-size:.72rem;color:#8b98a5'>💱 ₹{USD_TO_INR}/USD<br>📅 {date.today().strftime('%d %b %Y')}<br>🧾 <b style='color:#00a3ff'>{n}</b> invoice(s) saved<br>🔒 Session active</div>",
                    unsafe_allow_html=True)
        if st.button("Logout",use_container_width=True):
            for k in["authenticated","username","role","display_name"]:
                st.session_state.pop(k,None)
            st.rerun()
    return page


# ─── PAGE: DASHBOARD ─────────────────────────────────────────────────────────
def page_dashboard():
    st.markdown("<h2>📊 Financial Dashboard</h2>",unsafe_allow_html=True)
    inv=st.session_state["invoices"]   # FIX #4: direct reference, always live
    total_rev=sum(i["inr_amount"] for i in inv)
    total_gst=sum(i["total_gst"]  for i in inv)
    total_tds=sum(i["tds_amount"] for i in inv)
    total_net=sum(i["net_receivable"] for i in inv)
    c1,c2,c3,c4=st.columns(4)
    for col,lbl,val,sub in[(c1,"TOTAL REVENUE",fmt_inr(total_rev),f"{len(inv)} invoices"),
                           (c2,"GST LIABILITY",fmt_inr(total_gst),"CGST+SGST 18%"),
                           (c3,"TDS RECEIVABLE",fmt_inr(total_tds),"Claim in ITR"),
                           (c4,"NET RECEIVABLE",fmt_inr(total_net),"After TDS")]:
        with col: st.markdown(f"<div class='metric-card'><div class='metric-label'>{lbl}</div><div class='metric-value'>{val}</div><div class='metric-sub'>{sub}</div></div>",unsafe_allow_html=True)
    st.divider()
    ca,cb=st.columns(2)
    with ca:
        st.markdown("<div class='section-header'>Revenue Breakdown</div>",unsafe_allow_html=True)
        if inv:
            df=pd.DataFrame({"Inv":[f"#{i+1}" for i in range(len(inv))],
                             "Taxable":[x["inr_amount"] for x in inv],
                             "GST":[x["total_gst"] for x in inv],
                             "TDS":[x["tds_amount"] for x in inv]}).set_index("Inv")
            st.bar_chart(df,height=220)
        else: st.info("No invoices yet.")
    with cb:
        st.markdown("<div class='section-header'>Export vs Domestic</div>",unsafe_allow_html=True)
        ex=sum(i["inr_amount"] for i in inv if i.get("usd_amount"))
        dm=sum(i["inr_amount"] for i in inv if not i.get("usd_amount"))
        if ex+dm>0:
            st.bar_chart(pd.DataFrame({"Category":["Export","Domestic"],"Amount":[ex,dm]}).set_index("Category"),height=220)
        else: st.info("Will populate after invoices.")
    # FIX #3: Client-wise breakdown
    if inv:
        st.markdown("<div class='section-header'>👤 Client-Wise Summary</div>",unsafe_allow_html=True)
        cl={}
        for i in inv:
            cn=i.get("client_name","Unknown")
            if cn not in cl: cl[cn]={"Revenue":0,"GST":0,"TDS":0,"Net":0,"Count":0}
            cl[cn]["Revenue"]+=i["inr_amount"]; cl[cn]["GST"]+=i["total_gst"]
            cl[cn]["TDS"]+=i["tds_amount"];     cl[cn]["Net"]+=i["net_receivable"]
            cl[cn]["Count"]+=1
        st.dataframe(pd.DataFrame([{"Client":cn,"#":d["Count"],"Revenue":fmt_inr(d["Revenue"]),
                                     "GST":fmt_inr(d["GST"]),"TDS":fmt_inr(d["TDS"]),"Net":fmt_inr(d["Net"])}
                                    for cn,d in cl.items()]),use_container_width=True,hide_index=True)


# ─── PAGE: NEW INVOICE ───────────────────────────────────────────────────────
def page_new_invoice():
    st.markdown("<h2>🧾 Create Invoice</h2>",unsafe_allow_html=True)
    c1,c2=st.columns([1.2,1])
    with c1:
        st.markdown("<div class='section-header'>Invoice Details</div>",unsafe_allow_html=True)
        mode        =st.selectbox("Currency Mode",["USD (Export)","INR (Domestic)"])
        client_name =st.text_input("Client Name *",placeholder="e.g. Acme Corp")
        description =st.text_input("Work Description",placeholder="e.g. Web Dev – Q1 2026")
        amount      =st.number_input(f"Amount ({'USD' if 'USD' in mode else 'INR'})",
                                     min_value=0.0,value=1000.0,step=100.0,format="%.2f")
        tds_label   =st.selectbox("TDS Section",list(TDS_RATES.keys()))
        if st.button("⚡  Calculate & Save Invoice",use_container_width=True):
            if not client_name.strip():
                st.warning("Client Name is required.")
            else:
                i=calculate_invoice(mode,amount,tds_label,description,client_name)
                st.session_state["current_invoice"]=i
                save_invoice(i)
                st.success(f"✅ Saved! {len(st.session_state['invoices'])} total invoices in session.")
    with c2:
        st.markdown("<div class='section-header'>Live Preview</div>",unsafe_allow_html=True)
        i=st.session_state.get("current_invoice")
        if i:
            is_usd=i.get("usd_amount") is not None
            st.markdown("<span class='badge-export'>EXPORT / LUT</span>" if is_usd else "<span class='badge-inr'>DOMESTIC / GST</span>",unsafe_allow_html=True)
            rows=([("Client",i.get("client_name","—")),("USD Amount",fmt_usd(i["usd_amount"])),
                   ("Exchange Rate",f"₹{USD_TO_INR}"),("INR Equivalent",fmt_inr(i["inr_amount"])),
                   ("GST",i["gst_type"]),("TDS",fmt_inr(i["tds_amount"])),("✅ Net",fmt_inr(i["net_receivable"]))]
                  if is_usd else
                  [("Client",i.get("client_name","—")),("Taxable",fmt_inr(i["inr_amount"])),
                   ("CGST 9%",fmt_inr(i["cgst"])),("SGST 9%",fmt_inr(i["sgst"])),
                   ("Total GST",fmt_inr(i["total_gst"])),("Invoice Total",fmt_inr(i.get("invoice_total",0))),
                   ("TDS",fmt_inr(i["tds_amount"])),("✅ Net",fmt_inr(i["net_receivable"]))])
            st.dataframe(pd.DataFrame(rows,columns=["Field","Value"]),use_container_width=True,hide_index=True)
        else: st.info("Calculate an invoice to see preview.")


# ─── PAGE: HISTORY ───────────────────────────────────────────────────────────
def page_history():
    st.markdown("<h2>📁 Invoice History</h2>",unsafe_allow_html=True)
    inv=st.session_state["invoices"]   # FIX #4
    if not inv:
        st.info("No invoices yet.")
        return
    clients=sorted(set(i.get("client_name","Unknown") for i in inv))
    sel=st.selectbox("Filter by Client",["All"]+clients)
    filtered=inv if sel=="All" else [i for i in inv if i.get("client_name")==sel]
    rows=[{"#":idx+1,"Date":i["date"],"Client":i.get("client_name","—"),"Desc":i.get("description","—"),
           "Mode":i["mode"],"Taxable":i["inr_amount"],"GST":i["total_gst"],
           "TDS":i["tds_amount"],"Net":i["net_receivable"]} for idx,i in enumerate(filtered)]
    st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
    st.caption(f"{len(filtered)} of {len(inv)} invoices")
    if st.button("🗑  Clear All"):
        st.session_state["invoices"]=[]; st.session_state["current_invoice"]=None; st.rerun()


# ─── PAGE: AI ASSISTANT ──────────────────────────────────────────────────────
def page_ai():
    st.markdown("<h2>🤖 AI Financial Assistant</h2>",unsafe_allow_html=True)
    st.markdown("<div style='color:#8b98a5;font-size:.82rem'>Gemini 1.5 Flash · Multi-modal PDF/Image</div>",unsafe_allow_html=True)
    model=get_gemini()
    if not model:
        st.warning("⚠️ GEMINI_API_KEY missing in Secrets.")
        return
    tab1,tab2=st.tabs(["💬 Chat","📄 Scan Invoice → Dashboard"])

    with tab1:
        q=st.text_area("Ask about GST, TDS, ITR...",height=100,placeholder="e.g. Section 44ADA kya hai?")
        inv=st.session_state["invoices"]
        ctx=f"\n\nUser has {len(inv)} invoices. Revenue ₹{sum(i['inr_amount'] for i in inv):,.0f}, GST ₹{sum(i['total_gst'] for i in inv):,.0f}." if inv else ""
        if st.button("Ask Gemini ⚡"):
            if q.strip():
                with st.spinner("Consulting AI..."):
                    p=("You are DataSnap Pro AI advisor for Indian freelancers. "
                       "Give accurate answers on GST/TDS/ITR. Cite sections."+ctx+"\n\nQ: "+q)
                    r=model.generate_content(p)
                    st.markdown(f"<div class='ai-response'>{r.text}</div>",unsafe_allow_html=True)

    with tab2:
        st.markdown("<div class='section-header'>Upload Invoice → AI Extracts → Add to Dashboard</div>",unsafe_allow_html=True)
        uploaded=st.file_uploader("Upload PDF or Image",type=["pdf","png","jpg","jpeg"])
        if uploaded:
            fb=uploaded.read(); mime=uploaded.type
            st.success(f"✅ {uploaded.name} ({len(fb)//1024} KB)")
            if st.button("🔍 Extract with Gemini"):
                with st.spinner("Reading invoice..."):
                    b64=base64.standard_b64encode(fb).decode()
                    prompt="""Extract financial data from this invoice. Return ONLY valid JSON, no markdown:
{"client_name":"X","description":"Y","date":"DD Mon YYYY","mode":"USD (Export)","usd_amount":500.0,"inr_amount":46540.0,"cgst":0.0,"sgst":0.0,"total_gst":0.0,"tds_amount":0.0,"net_receivable":46540.0}
For INR domestic: set usd_amount to null. Fill all fields accurately."""
                    resp=model.generate_content([{"inline_data":{"mime_type":mime,"data":b64}},{"text":prompt}])
                    raw=resp.text.strip().replace("```json","").replace("```python","").replace("```","").strip()
                    st.markdown(f"<div class='ai-response'>{raw}</div>",unsafe_allow_html=True)
                    # FIX #2: store parsed result, bridge button rendered OUTSIDE this block
                    try:
                        parsed=json.loads(raw)
                        parsed.setdefault("mode","INR – Domestic" if not parsed.get("usd_amount") else "USD – Export of Service")
                        parsed.setdefault("gst_type","LUT / Export – 0% GST" if parsed.get("usd_amount") else "18% GST (CGST 9% + SGST 9%)")
                        parsed.setdefault("tds_rate",0.0)
                        parsed.setdefault("exchange_rate",USD_TO_INR if parsed.get("usd_amount") else None)
                        parsed.setdefault("invoice_total",parsed.get("inr_amount",0)+parsed.get("total_gst",0))
                        parsed.setdefault("saved_by",st.session_state.get("username","ai"))
                        st.session_state["ai_extracted"]=parsed
                    except Exception as e:
                        st.session_state["ai_extracted"]=None
                        st.warning(f"Auto-parse failed ({e}). Add manually via New Invoice.")

        # FIX #2: Bridge button is OUTSIDE the upload/analyse block — persists across reruns
        ex=st.session_state.get("ai_extracted")
        if ex:
            st.markdown("<div class='bridge-card'>",unsafe_allow_html=True)
            st.markdown("**📋 Extracted — confirm to add:**")
            preview=[(k,str(v)) for k,v in ex.items() if k not in["gst_type","tds_rate","exchange_rate","saved_by","invoice_total"]]
            st.dataframe(pd.DataFrame(preview,columns=["Field","Value"]),use_container_width=True,hide_index=True)
            col_a,col_b=st.columns(2)
            with col_a:
                if st.button("📥 Add to Dashboard",use_container_width=True):
                    save_invoice(ex)
                    st.session_state["ai_extracted"]=None
                    st.success(f"✅ Added! Total: {len(st.session_state['invoices'])} invoices")
                    st.rerun()
            with col_b:
                if st.button("🗑 Discard",use_container_width=True):
                    st.session_state["ai_extracted"]=None; st.rerun()
            st.markdown("</div>",unsafe_allow_html=True)


# ─── PAGE: EXPORT ────────────────────────────────────────────────────────────
def page_export():
    st.markdown("<h2>📤 Export & Notifications</h2>",unsafe_allow_html=True)
    inv=st.session_state["invoices"]   # FIX #4 + #5
    tab1,tab2=st.tabs(["📊 Excel Report","💬 WhatsApp"])

    with tab1:
        st.markdown("<div class='section-header'>CA Audit Excel Report</div>",unsafe_allow_html=True)
        if not inv:
            st.info("No invoices yet — create some first.")
        else:
            st.success(f"✅ {len(inv)} invoice(s) ready to export.")
            owner=st.text_input("Freelancer / Owner Name",value=st.session_state.get("display_name",""))
            # FIX #5: bytes generated immediately; download_button is NOT inside another button
            xls=build_excel(inv, owner or "—")
            st.download_button(
                label=f"📥 Download Excel — {len(inv)} invoices",
                data=xls,
                file_name=f"DataSnap_Audit_{date.today().isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.caption("Includes: CA Audit · Client Summary · Overall Summary — 3 sheets")

    with tab2:
        st.markdown("<div class='section-header'>WhatsApp Notification</div>",unsafe_allow_html=True)
        st.markdown("<div style='background:#0a1833;border:1px solid #1e2d3d;border-radius:8px;padding:.9rem 1.1rem;margin-bottom:1rem;font-size:.78rem;color:#8b98a5'>Pre-wired for <b style='color:#00a3ff'>Twilio/WATI</b>. Uncomment code in <code>send_whatsapp_placeholder()</code> and add Secrets.</div>",unsafe_allow_html=True)
        phone=st.text_input("WhatsApp Number",placeholder="+91XXXXXXXXXX")
        c_name=st.text_input("Client Name",value="Client")
        sel_inv=st.session_state.get("current_invoice") or (inv[-1] if inv else None)
        if sel_inv:
            msg=whatsapp_message(sel_inv,c_name,phone)
            st.markdown("**Preview:**")
            st.markdown(f"<div class='whatsapp-preview'>{msg}</div>",unsafe_allow_html=True)
            if st.button("📲 Send (Mock)"):
                r=send_whatsapp_placeholder(phone,msg)
                st.success(f"Mock → {r['to']} | {r['chars']} chars | {r['status']}")
        else:
            st.info("Create an invoice first.")


# ─── PAGE: ADMIN ─────────────────────────────────────────────────────────────
def page_admin():
    st.markdown("<h2>👥 Admin Panel</h2>",unsafe_allow_html=True)
    st.markdown("<div class='section-header'>Users</div>",unsafe_allow_html=True)
    st.dataframe(pd.DataFrame([{"Username":u,"Name":d["name"],"Role":d["role"]} for u,d in USERS.items()]),
                 use_container_width=True,hide_index=True)
    st.info("To add users: edit USERS dict or connect SQLite / Supabase.")
    st.markdown("<div class='section-header'>Live Config</div>",unsafe_allow_html=True)
    st.json({"USD_TO_INR":USD_TO_INR,"CGST":"9%","SGST":"9%","GST":"18%",
             "Model":"gemini-1.5-flash","GSheets":"stub (uncomment to activate)",
             "Total_Invoices_This_Session":len(st.session_state["invoices"])})


# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    inject_css()
    # init_state() already ran at module top
    if not st.session_state["authenticated"]:
        login_block(); return
    page=render_sidebar()
    if   "Dashboard" in page: page_dashboard()
    elif "Invoice"   in page: page_new_invoice()
    elif "History"   in page: page_history()
    elif "AI"        in page: page_ai()
    elif "Export"    in page: page_export()
    elif "Admin"     in page: page_admin()

if __name__=="__main__":
    main()
