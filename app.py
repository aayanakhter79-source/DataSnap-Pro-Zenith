"""
DataSnap Pro — AI-Powered Financial OS for Freelancers
Built by Zenith IN | Version 3.0.0
v3 Changes:
  1. AI ↔ Manual Bridge  — Gemini data flows through same calculate_invoice() logic
  2. Dynamic Exchange Rate — Sidebar number_input drives ALL calculations
  3. Currency Selector     — Sidebar + AI-aware currency detection
  4. NoneType Safety       — All numeric fields null-guarded before any arithmetic
  5. Clean AI UI           — No raw JSON; structured st.dataframe preview
"""

import streamlit as st
import pandas as pd
import io, hashlib, base64, json
from datetime import datetime, date
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════════════════════════════
#  PAGE CONFIG  — must be the very first Streamlit call
# ════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="DataSnap Pro | Zenith IN",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ════════════════════════════════════════════════════════════════════════════
#  SESSION STATE INIT  — safe across every rerun
# ════════════════════════════════════════════════════════════════════════════
_DEFAULTS = {
    "invoices": [],
    "current_invoice": None,
    "ai_extracted": None,  # holds parsed AI dict until user confirms
    "authenticated": False,
    "username": "",
    "role": "Client",
    "display_name": "User",
    # v3: live rate stored in session so it survives navigation
    "usd_rate": 93.08,
    "default_currency": "USD (Export)",
    # v4 new features
    "expenses": [],  # Expense tracker entries
    "rag_chat_history": [],  # RAG chat memory
}
for _k, _v in _DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ════════════════════════════════════════════════════════════════════════════
#  CONSTANTS  (only the ones that never change)
# ════════════════════════════════════════════════════════════════════════════
CGST_RATE = 0.09
SGST_RATE = 0.09
TDS_RATES = {
    "194J – Professional / Technical (10%)": 0.10,
    "194J – Royalty / FTS (2%)": 0.02,
    "194C – Contractor Individual/HUF (1%)": 0.01,
    "194C – Contractor Company (2%)": 0.02,
    "No TDS": 0.00,
}

EXPENSE_CATEGORIES = [
    "Software / Tools",
    "Internet / Phone",
    "Co-working Space",
    "Hardware / Equipment",
    "Learning / Courses",
    "Marketing / Ads",
    "Travel / Transport",
    "Office Supplies",
    "Contractor / Subcontractor",
    "Bank Charges",
    "Other",
]

# ════════════════════════════════════════════════════════════════════════════
#  USERS
# ════════════════════════════════════════════════════════════════════════════
USERS = {
    "admin": {
        "password": hashlib.sha256(b"zenith@2026").hexdigest(),
        "role": "Admin",
        "name": "Admin User",
    },
    "client1": {
        "password": hashlib.sha256(b"client1pass").hexdigest(),
        "role": "Client",
        "name": "Rahul Sharma",
    },
    "client2": {
        "password": hashlib.sha256(b"client2pass").hexdigest(),
        "role": "Client",
        "name": "Priya Mehta",
    },
}

# ════════════════════════════════════════════════════════════════════════════
#  GOOGLE SHEETS STUB  (uncomment when ready)
# ════════════════════════════════════════════════════════════════════════════
# import gspread
# from google.oauth2.service_account import Credentials
#
# @st.cache_resource
# def get_gsheet():
#     creds = Credentials.from_service_account_info(
#         st.secrets["gcp_service_account"],
#         scopes=["https://www.googleapis.com/auth/spreadsheets"])
#     return gspread.authorize(creds).open_by_key(st.secrets["SHEET_ID"]).sheet1
#
# def _push_to_gsheet(inv: dict):
#     get_gsheet().append_row([
#         inv["date"], inv.get("client_name",""), inv.get("description",""),
#         inv["mode"], inv.get("usd_amount",""), inv["inr_amount"],
#         inv["cgst"], inv["sgst"], inv["total_gst"],
#         inv.get("invoice_total", inv["inr_amount"]),
#         inv["tds_amount"], inv["net_receivable"],
#     ])


def _push_to_gsheet(inv: dict):
    pass  # STUB — replace with real impl above when ready


def save_invoice(inv: dict):
    """Single entry point for persisting an invoice — session + optional cloud."""
    st.session_state["invoices"].append(inv)
    _push_to_gsheet(inv)


# ════════════════════════════════════════════════════════════════════════════
#  GEMINI
# ════════════════════════════════════════════════════════════════════════════
@st.cache_resource
def get_gemini():
    try:
        genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
        available = [
            m.name
            for m in genai.list_models()
            if "generateContent" in m.supported_generation_methods
        ]
        name = (
            "models/gemini-1.5-flash"
            if "models/gemini-1.5-flash" in available
            else (available[0] if available else None)
        )
        return genai.GenerativeModel(name) if name else None
    except Exception:
        return None


# ════════════════════════════════════════════════════════════════════════════
#  CSS — Zenith Dark Theme
# ════════════════════════════════════════════════════════════════════════════
def inject_css():
    st.markdown(
        """
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@300;400;500&display=swap');

:root {
    --black:#080c10; --surface:#0d1117; --card:#111820; --border:#1e2d3d;
    --blue:#00a3ff;  --blue2:#0066cc;   --white:#f0f6fc; --muted:#8b98a5;
    --green:#00e676; --red:#ff5252;     --gold:#ffd700;  --teal:#00d4aa;
}
html,body,[data-testid="stAppViewContainer"]{
    background-color:var(--black)!important;
    color:var(--white)!important;
    font-family:'DM Mono',monospace!important;
}
[data-testid="stSidebar"]{
    background:var(--surface)!important;
    border-right:1px solid var(--border)!important;
}
[data-testid="stSidebar"] *{color:var(--white)!important;}
h1,h2,h3{font-family:'Syne',sans-serif!important;font-weight:800!important;}

/* ── Metric Cards ── */
.metric-card{
    background:var(--card);border:1px solid var(--border);
    border-top:3px solid var(--blue);border-radius:8px;
    padding:1.2rem 1.4rem;margin-bottom:.8rem;
}
.metric-label{font-size:.72rem;letter-spacing:.12em;color:var(--muted);text-transform:uppercase;}
.metric-value{font-size:1.9rem;font-weight:700;font-family:'Syne',sans-serif;color:var(--white);line-height:1.2;}
.metric-sub{font-size:.75rem;color:var(--blue);margin-top:.2rem;}

/* ── Rate badge in sidebar ── */
.rate-chip{
    background:#001a2e;border:1px solid var(--blue);border-radius:20px;
    padding:4px 12px;font-size:.75rem;color:var(--blue);display:inline-block;
    font-weight:700;letter-spacing:.05em;margin-top:.3rem;
}
.currency-chip-usd{
    background:#00332a;border:1px solid var(--green);border-radius:20px;
    padding:3px 10px;font-size:.7rem;color:var(--green);font-weight:700;
}
.currency-chip-inr{
    background:#2a1a00;border:1px solid var(--gold);border-radius:20px;
    padding:3px 10px;font-size:.7rem;color:var(--gold);font-weight:700;
}

/* ── Badges ── */
.badge-export{background:#00332a;color:var(--green);padding:2px 8px;border-radius:4px;font-size:.7rem;font-weight:600;border:1px solid var(--green);}
.badge-inr{background:#2a1a00;color:var(--gold);padding:2px 8px;border-radius:4px;font-size:.7rem;font-weight:600;border:1px solid var(--gold);}
.badge-admin{background:#001833;color:var(--blue);padding:2px 8px;border-radius:4px;font-size:.7rem;font-weight:600;border:1px solid var(--blue);}
.badge-ai{background:#0a0a2a;color:var(--teal);padding:2px 8px;border-radius:4px;font-size:.7rem;font-weight:600;border:1px solid var(--teal);}

/* ── Section header ── */
.section-header{
    border-left:3px solid var(--blue);padding-left:.8rem;
    margin:1.5rem 0 1rem;font-family:'Syne',sans-serif;
    font-size:1.1rem;font-weight:700;color:var(--white);
}

/* ── Buttons ── */
.stButton>button{
    background:linear-gradient(135deg,var(--blue2),var(--blue))!important;
    color:#fff!important;border:none!important;border-radius:6px!important;
    font-family:'Syne',sans-serif!important;font-weight:700!important;
    letter-spacing:.05em!important;padding:.55rem 1.4rem!important;
    transition:opacity .2s!important;
}
.stButton>button:hover{opacity:.85!important;}

/* ── Inputs ── */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input,
.stTextArea textarea,
[data-testid="stSelectbox"] > div > div{
    background:var(--card)!important;color:var(--white)!important;
    border:1px solid var(--border)!important;border-radius:6px!important;
    font-family:'DM Mono',monospace!important;
}
[data-testid="stDataFrame"]{background:var(--card)!important;}

/* ── Logo ── */
.zenith-logo{font-family:'Syne',sans-serif;font-size:1.4rem;font-weight:800;letter-spacing:.08em;color:var(--blue);}
.zenith-sub{font-size:.68rem;color:var(--muted);letter-spacing:.15em;text-transform:uppercase;}

/* ── WhatsApp preview ── */
.whatsapp-preview{
    background:#0a1a0f;border:1px solid #1a4a1a;border-radius:8px;
    padding:1rem 1.2rem;font-size:.82rem;color:#cce8cc;
    white-space:pre-wrap;font-family:'DM Mono',monospace;
}

/* ── AI response ── */
.ai-response{
    background:var(--card);border:1px solid var(--border);
    border-left:3px solid var(--blue);border-radius:8px;
    padding:1rem 1.2rem;font-size:.85rem;line-height:1.6;color:var(--white);
}

/* ── AI bridge card ── */
.bridge-card{
    background:#001020;border:1px solid #003366;border-radius:10px;
    padding:1.2rem 1.4rem;margin-top:1rem;
}
.bridge-title{
    font-family:'Syne',sans-serif;font-size:1rem;font-weight:700;
    color:var(--teal);margin-bottom:.8rem;
    border-bottom:1px solid #003366;padding-bottom:.5rem;
}

/* ── Info / warning box ── */
.info-box{
    background:#0a1833;border:1px solid #1e2d3d;border-radius:8px;
    padding:.9rem 1.1rem;margin-bottom:1rem;font-size:.78rem;color:#8b98a5;
}

.stTabs [data-baseweb="tab"]{
    background:transparent!important;color:var(--muted)!important;
    font-family:'Syne',sans-serif!important;font-weight:600!important;
    border-radius:0!important;border-bottom:2px solid transparent!important;
}
.stTabs [aria-selected="true"]{
    color:var(--blue)!important;
    border-bottom:2px solid var(--blue)!important;
    background:transparent!important;
}
div[data-testid="stAlert"]{border-radius:6px!important;}
footer{display:none!important;}

/* ── v4: Expense & Tax cards ── */
.expense-tag{
    display:inline-block;padding:2px 8px;border-radius:4px;
    font-size:.68rem;font-weight:700;letter-spacing:.04em;
    background:#1a0a00;color:var(--gold);border:1px solid var(--gold);
}
.profit-card{
    background:var(--card);border:1px solid var(--border);
    border-top:3px solid var(--green);border-radius:8px;
    padding:1.2rem 1.4rem;margin-bottom:.8rem;
}
.loss-card{
    background:var(--card);border:1px solid var(--border);
    border-top:3px solid var(--red);border-radius:8px;
    padding:1.2rem 1.4rem;margin-bottom:.8rem;
}
.tax-pill{
    display:inline-block;padding:4px 12px;border-radius:20px;
    font-size:.75rem;font-weight:700;letter-spacing:.05em;
    background:#001a0a;color:var(--green);border:1px solid var(--green);
}
.rag-msg-user{
    background:#001020;border-left:3px solid var(--blue);border-radius:6px;
    padding:.7rem 1rem;margin:.5rem 0;font-size:.83rem;color:var(--white);
}
.rag-msg-ai{
    background:#0a1a0a;border-left:3px solid var(--green);border-radius:6px;
    padding:.7rem 1rem;margin:.5rem 0;font-size:.83rem;color:var(--white);
    line-height:1.6;
}
.rag-source{
    font-size:.68rem;color:var(--muted);margin-top:.4rem;
    border-top:1px solid var(--border);padding-top:.3rem;
}
</style>
""",
        unsafe_allow_html=True,
    )


# ════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ════════════════════════════════════════════════════════════════════════════
fmt_inr = lambda v: f"₹{float(v or 0):,.2f}"
fmt_usd = lambda v: f"${float(v or 0):,.2f}"
hash_pw = lambda pw: hashlib.sha256(pw.encode()).hexdigest()


def live_rate() -> float:
    """Always read the live USD rate from session state."""
    return float(st.session_state.get("usd_rate", 93.08))


# FIX #4 — null-guard: convert None / missing / non-numeric to 0.0
def safe_float(val, default: float = 0.0) -> float:
    try:
        if val is None or val == "" or val == "null":
            return default
        return float(val)
    except (TypeError, ValueError):
        return default


# ════════════════════════════════════════════════════════════════════════════
#  CORE INVOICE CALCULATOR
#  Used by BOTH manual entry AND AI bridge — single source of truth
# ════════════════════════════════════════════════════════════════════════════
def calculate_invoice(
    mode: str,
    amount: float,
    tds_label: str,
    description: str = "",
    client_name: str = "",
    override_rate: float | None = None,  # AI can pass extracted rate
) -> dict:
    """
    Build a fully-formed invoice dict.
    mode: 'USD (Export)' | 'INR (Domestic)'
    All callers (manual + AI bridge) use this function — guaranteed consistency.
    """
    rate = override_rate if override_rate else live_rate()
    tds_rate = TDS_RATES.get(tds_label, 0.0)
    amount = safe_float(amount)

    if mode == "USD (Export)":
        inr = amount * rate
        inv = dict(
            mode="USD – Export of Service",
            usd_amount=amount,
            inr_amount=inr,
            gst_type="LUT / Export – 0% GST",
            cgst=0.0,
            sgst=0.0,
            total_gst=0.0,
            taxable_value=inr,
            invoice_total=inr,
            tds_rate=tds_rate,
            tds_amount=inr * tds_rate,
            net_receivable=inr * (1 - tds_rate),
            exchange_rate=rate,
        )
    else:  # INR Domestic
        cgst = amount * CGST_RATE
        sgst = amount * SGST_RATE
        total_gst = cgst + sgst
        inv_total = amount + total_gst
        tds_amt = amount * tds_rate
        inv = dict(
            mode="INR – Domestic",
            usd_amount=None,
            inr_amount=amount,
            gst_type="18% GST (CGST 9% + SGST 9%)",
            cgst=cgst,
            sgst=sgst,
            total_gst=total_gst,
            taxable_value=amount,
            invoice_total=inv_total,
            tds_rate=tds_rate,
            tds_amount=tds_amt,
            net_receivable=inv_total - tds_amt,
            exchange_rate=None,
        )

    inv.update(
        description=description.strip(),
        client_name=(
            client_name.strip() or description.split("–")[0].strip() or "Unknown"
        ),
        date=datetime.now().strftime("%d %b %Y"),
        saved_by=st.session_state.get("username", ""),
    )
    return inv


# ════════════════════════════════════════════════════════════════════════════
#  AI EXTRACTION → NORMALISE
#  Converts raw Gemini JSON into the same shape calculate_invoice() produces
# ════════════════════════════════════════════════════════════════════════════
def normalise_ai_extraction(raw: dict) -> dict:
    """
    Run the AI-parsed dict through calculate_invoice() so ALL fields
    are consistent, null-safe, and use the live exchange rate.
    """
    # Detect currency from AI output
    usd_raw = safe_float(raw.get("usd_amount"))
    inr_raw = safe_float(raw.get("inr_amount"))
    ai_mode_hint = str(raw.get("mode", "")).lower()

    if usd_raw > 0 or "export" in ai_mode_hint or "usd" in ai_mode_hint:
        mode = "USD (Export)"
        amount = usd_raw if usd_raw > 0 else (inr_raw / live_rate() if inr_raw else 0.0)
    else:
        mode = "INR (Domestic)"
        amount = inr_raw

    # Try to recover TDS label from rate
    tds_rate_raw = safe_float(raw.get("tds_rate"))
    tds_label = "No TDS"
    for label, rate in TDS_RATES.items():
        if abs(rate - tds_rate_raw) < 0.001:
            tds_label = label
            break

    # Re-run through the canonical calculator
    normalised = calculate_invoice(
        mode=mode,
        amount=amount,
        tds_label=tds_label,
        description=str(raw.get("description", "")).strip(),
        client_name=str(raw.get("client_name", "")).strip(),
    )
    # Preserve AI-detected date if available
    ai_date = str(raw.get("date", "")).strip()
    if ai_date and ai_date != "null":
        normalised["date"] = ai_date

    return normalised


# ════════════════════════════════════════════════════════════════════════════
#  WHATSAPP
# ════════════════════════════════════════════════════════════════════════════
def whatsapp_message(inv: dict, client_name: str, phone: str) -> str:
    lines = [
        f"👋 Hello {client_name}!",
        "",
        "📋 *Invoice Summary — DataSnap Pro*",
        f"🗓  Date : {inv['date']}",
        f"📝 Desc : {inv.get('description', '—')}",
        "",
        f"💰 Mode : {inv['mode']}",
    ]
    if inv.get("usd_amount"):
        lines.append(
            f"   USD   : {fmt_usd(inv['usd_amount'])} → {fmt_inr(inv['inr_amount'])}"
        )
        lines.append(f"   Rate  : ₹{inv.get('exchange_rate', live_rate())}/USD")
    else:
        lines += [
            f"   Taxable    : {fmt_inr(inv['inr_amount'])}",
            f"   GST (18%)  : {fmt_inr(inv['total_gst'])}",
            f"   Inv Total  : {fmt_inr(inv.get('invoice_total', inv['inr_amount']))}",
        ]
    if inv["tds_amount"] > 0:
        lines.append(
            f"   TDS ({inv['tds_rate']*100:.0f}%)  : {fmt_inr(inv['tds_amount'])}"
        )
    lines += [
        "",
        f"✅ *Net Receivable : {fmt_inr(inv['net_receivable'])}*",
        "",
        "Powered by ⚡ Zenith IN / DataSnap Pro",
    ]
    return "\n".join(lines)


def send_whatsapp_placeholder(phone: str, message: str) -> dict:
    # Twilio: Client(sid,tok).messages.create(from_='whatsapp:+14155238886',to=f'whatsapp:{phone}',body=message)
    # WATI  : requests.post(url,headers={"Authorization":f"Bearer {tok}"},json={"messageText":message})
    return {"status": "mock_sent", "to": phone, "chars": len(message)}


# ════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ════════════════════════════════════════════════════════════════════════════
def build_excel(invoices: list, owner: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "CA Audit Report"
    D = "0D1117"
    H = "0A2540"
    B = "00A3FF"
    W = "F0F6FC"
    M = "8B98A5"
    G = "00E676"
    GO = "FFD700"
    R = "FF5252"

    def cs(sh, r, c, val, bold=False, fg=W, bg=D, al="left", nf=None, sz=10):
        cell = sh.cell(row=r, column=c, value=val)
        cell.font = Font(name="Calibri", bold=bold, color=fg, size=sz)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=al, vertical="center", wrap_text=True)
        cell.border = Border(
            **{
                s: Side(style="thin", color="1E2D3D")
                for s in ["left", "right", "top", "bottom"]
            }
        )
        if nf:
            cell.number_format = nf
        return cell

    # Title
    ws.merge_cells("A1:M1")
    cs(
        ws,
        1,
        1,
        f"⚡  DataSnap Pro  |  CA Audit Report  |  Zenith IN",
        bold=True,
        fg=B,
        bg="080C10",
        al="center",
        sz=14,
    )
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:M2")
    cs(
        ws,
        2,
        1,
        f"For: {owner}   |   {datetime.now().strftime('%d %b %Y %H:%M')}   |   Rate: ₹{live_rate()}/USD",
        fg=M,
        bg="080C10",
        al="center",
        sz=9,
    )
    ws.row_dimensions[2].height = 18

    hdrs = [
        "#",
        "Date",
        "Client",
        "Description",
        "Mode",
        "USD",
        "INR (Taxable)",
        "CGST 9%",
        "SGST 9%",
        "GST Total",
        "Invoice Total",
        "TDS",
        "Net Receivable",
    ]
    wds = [5, 14, 18, 30, 20, 14, 16, 12, 12, 12, 14, 12, 16]
    for i, (h, w) in enumerate(zip(hdrs, wds), 1):
        cs(ws, 3, i, h, bold=True, fg=B, bg=H, al="center", sz=10)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 22

    tot = dict(usd=0, inr=0, cgst=0, sgst=0, gst=0, ivt=0, tds=0, net=0)
    for idx, inv in enumerate(invoices, 1):
        r = idx + 3
        bg2 = D if idx % 2 == 0 else "0F161F"
        is_usd = inv.get("usd_amount") not in [None, 0, 0.0]
        ivt = safe_float(inv.get("invoice_total", inv.get("inr_amount")))
        vals = [
            idx,
            inv["date"],
            inv.get("client_name", "—"),
            inv.get("description", "—"),
            inv["mode"],
            safe_float(inv.get("usd_amount")) if is_usd else "—",
            safe_float(inv.get("inr_amount")),
            safe_float(inv.get("cgst")),
            safe_float(inv.get("sgst")),
            safe_float(inv.get("total_gst")),
            ivt,
            safe_float(inv.get("tds_amount")),
            safe_float(inv.get("net_receivable")),
        ]
        fmts = [
            None,
            None,
            None,
            None,
            None,
            "[$$-en-US]#,##0.00",
            "₹#,##0.00",
            "₹#,##0.00",
            "₹#,##0.00",
            "₹#,##0.00",
            "₹#,##0.00",
            "₹#,##0.00",
            "₹#,##0.00",
        ]
        fgs = [M, M, W, W, (GO if is_usd else G), W, W, W, W, W, W, R, G]
        for ci, (v, nf, fg) in enumerate(zip(vals, fmts, fgs), 1):
            cs(ws, r, ci, v, fg=fg, bg=bg2, nf=nf, al="right" if ci > 5 else "left")
        tot["usd"] += safe_float(inv.get("usd_amount"))
        tot["inr"] += safe_float(inv.get("inr_amount"))
        tot["cgst"] += safe_float(inv.get("cgst"))
        tot["sgst"] += safe_float(inv.get("sgst"))
        tot["gst"] += safe_float(inv.get("total_gst"))
        tot["ivt"] += ivt
        tot["tds"] += safe_float(inv.get("tds_amount"))
        tot["net"] += safe_float(inv.get("net_receivable"))

    tr = len(invoices) + 4
    ws.row_dimensions[tr].height = 22
    tv = [
        "",
        "TOTALS",
        "",
        "",
        "",
        tot["usd"],
        tot["inr"],
        tot["cgst"],
        tot["sgst"],
        tot["gst"],
        tot["ivt"],
        tot["tds"],
        tot["net"],
    ]
    tf = [
        None,
        None,
        None,
        None,
        None,
        "[$$-en-US]#,##0.00",
        "₹#,##0.00",
        "₹#,##0.00",
        "₹#,##0.00",
        "₹#,##0.00",
        "₹#,##0.00",
        "₹#,##0.00",
        "₹#,##0.00",
    ]
    for ci, (v, nf) in enumerate(zip(tv, tf), 1):
        cs(
            ws,
            tr,
            ci,
            v,
            bold=True,
            fg=B,
            bg=H,
            nf=nf,
            al="right" if ci > 5 else "left",
            sz=11,
        )

    # ── Client Summary sheet ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Client Summary")
    ws3.sheet_view.showGridLines = False
    cl = {}
    for inv in invoices:
        cn = inv.get("client_name", "Unknown")
        if cn not in cl:
            cl[cn] = {"rev": 0, "gst": 0, "tds": 0, "net": 0, "n": 0}
        cl[cn]["rev"] += safe_float(inv.get("inr_amount"))
        cl[cn]["gst"] += safe_float(inv.get("total_gst"))
        cl[cn]["tds"] += safe_float(inv.get("tds_amount"))
        cl[cn]["net"] += safe_float(inv.get("net_receivable"))
        cl[cn]["n"] += 1
    for i, (h, w) in enumerate(
        zip(
            ["Client", "Invoices", "Revenue", "GST", "TDS", "Net"],
            [25, 10, 18, 14, 14, 18],
        ),
        1,
    ):
        cs(ws3, 1, i, h, bold=True, fg=B, bg=H, al="center")
        ws3.column_dimensions[get_column_letter(i)].width = w
    for ri, (cn, d) in enumerate(cl.items(), 2):
        bg3 = D if ri % 2 == 0 else "0F161F"
        for ci, (v, nf) in enumerate(
            zip(
                [cn, d["n"], d["rev"], d["gst"], d["tds"], d["net"]],
                [None, None, "₹#,##0.00", "₹#,##0.00", "₹#,##0.00", "₹#,##0.00"],
            ),
            1,
        ):
            cs(ws3, ri, ci, v, fg=W, bg=bg3, nf=nf, al="right" if ci > 1 else "left")
        ws3.row_dimensions[ri].height = 20

    # ── Overall Summary sheet ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    sd = [
        ("METRIC", "AMOUNT", "NOTES"),
        ("Total Invoices", len(invoices), "Count"),
        ("Total Revenue (INR)", tot["inr"], "Taxable value"),
        ("Total GST Liability", tot["gst"], "CGST + SGST"),
        ("  ↳ CGST (9%)", tot["cgst"], "Central GST"),
        ("  ↳ SGST (9%)", tot["sgst"], "State GST"),
        ("TDS Receivable", tot["tds"], "Claim in ITR"),
        ("Net Receivable", tot["net"], "After TDS"),
        ("USD Revenue", tot["usd"], "Export of Service (LUT)"),
    ]
    for ri, row in enumerate(sd, 1):
        ih = ri == 1
        bg4 = H if ih else (D if ri % 2 == 0 else "0F161F")
        for ci, val in enumerate(row, 1):
            fg4 = B if ih else (G if ci == 2 and ri > 1 else W)
            nf4 = (
                "₹#,##0.00"
                if ci == 2 and ri not in [1, 2, 9]
                else ("[$$-en-US]#,##0.00" if ci == 2 and ri == 9 else None)
            )
            cs(
                ws2,
                ri,
                ci,
                val,
                bold=ih,
                fg=fg4,
                bg=bg4,
                nf=nf4,
                al="right" if ci == 2 else "left",
            )
        ws2.row_dimensions[ri].height = 20
    for ci, w in [(1, 30), (2, 20), (3, 30)]:
        ws2.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════
#  LOGIN
# ════════════════════════════════════════════════════════════════════════════
def login_block():
    st.markdown(
        """
<div style='text-align:center;padding:3rem 0 1.5rem'>
    <div class='zenith-logo'>⚡ ZENITH IN</div>
    <div class='zenith-sub'>DataSnap Pro · Financial OS v3.0</div>
</div>""",
        unsafe_allow_html=True,
    )
    col = st.columns([1, 1.2, 1])[1]
    with col:
        st.markdown("#### Sign In")
        u = st.text_input("Username", key="li_user")
        p = st.text_input("Password", type="password", key="li_pw")
        if st.button("LOGIN  →", use_container_width=True):
            usr = USERS.get(u)
            if usr and usr["password"] == hash_pw(p):
                st.session_state.update(
                    {
                        "authenticated": True,
                        "username": u,
                        "role": usr["role"],
                        "display_name": usr["name"],
                    }
                )
                st.rerun()
            else:
                st.error("Invalid credentials.")
        st.markdown(
            """
<div style='font-size:.72rem;color:#8b98a5;margin-top:1rem'>
Demo → admin / zenith@2026<br>
Demo → client1 / client1pass
</div>""",
            unsafe_allow_html=True,
        )


# ════════════════════════════════════════════════════════════════════════════
#  SIDEBAR  — v3: includes live rate input + default currency selector
# ════════════════════════════════════════════════════════════════════════════
def render_sidebar() -> str:
    with st.sidebar:
        st.markdown(
            """
<div style='padding:1rem 0 .5rem'>
    <div class='zenith-logo'>⚡ ZENITH IN</div>
    <div class='zenith-sub'>DataSnap Pro v3.0</div>
</div>""",
            unsafe_allow_html=True,
        )
        st.divider()

        # ── User info ────────────────────────────────────────────────────────
        name = st.session_state["display_name"]
        role = st.session_state["role"]
        badge = "badge-admin" if role == "Admin" else "badge-inr"
        st.markdown(
            f"""
<div style='margin-bottom:1rem'>
    <div style='font-size:.8rem;color:#8b98a5'>Logged in as</div>
    <div style='font-size:1rem;font-weight:700'>{name}</div>
    <span class='{badge}'>{role}</span>
</div>""",
            unsafe_allow_html=True,
        )

        # ── v3: DYNAMIC EXCHANGE RATE ─────────────────────────────────────────
        st.markdown(
            "<div style='font-size:.75rem;color:#8b98a5;letter-spacing:.08em;text-transform:uppercase;margin-bottom:.3rem'>💱 Live USD Rate</div>",
            unsafe_allow_html=True,
        )
        new_rate = st.number_input(
            "USD → INR Rate",
            min_value=50.0,
            max_value=200.0,
            value=float(st.session_state["usd_rate"]),
            step=0.01,
            format="%.2f",
            label_visibility="collapsed",
            key="rate_input",
        )
        st.session_state["usd_rate"] = new_rate
        st.markdown(
            f"<div class='rate-chip'>₹{new_rate:.2f} / USD</div>",
            unsafe_allow_html=True,
        )

        st.markdown("<br>", unsafe_allow_html=True)

        # ── v3: DEFAULT CURRENCY SELECTOR ────────────────────────────────────
        st.markdown(
            "<div style='font-size:.75rem;color:#8b98a5;letter-spacing:.08em;text-transform:uppercase;margin-bottom:.3rem'>🌐 Default Mode</div>",
            unsafe_allow_html=True,
        )
        curr_idx = 0 if st.session_state["default_currency"] == "USD (Export)" else 1
        chosen_currency = st.radio(
            "Default currency mode",
            ["USD (Export)", "INR (Domestic)"],
            index=curr_idx,
            label_visibility="collapsed",
            key="currency_radio",
        )
        st.session_state["default_currency"] = chosen_currency
        chip_cls = (
            "currency-chip-usd" if "USD" in chosen_currency else "currency-chip-inr"
        )
        chip_lbl = (
            "🟢 USD / Export" if "USD" in chosen_currency else "🟡 INR / Domestic"
        )
        st.markdown(
            f"<div class='{chip_cls}' style='margin-top:.3rem'>{chip_lbl}</div>",
            unsafe_allow_html=True,
        )

        st.divider()

        # ── Navigation ───────────────────────────────────────────────────────
        pages = [
            "📊 Dashboard",
            "🧾 New Invoice",
            "📁 Invoice History",
            "🤖 AI Assistant",
            "🧠 Smart Data Assistant",
            "💸 Expense Tracker",
            "🧮 Tax Planner",
            "📤 Export & WhatsApp",
        ]
        if role == "Admin":
            pages.append("👥 Admin Panel")
        page = st.radio("Navigation", pages, label_visibility="collapsed")

        st.divider()
        n = len(st.session_state["invoices"])
        ne = len(st.session_state["expenses"])
        st.markdown(
            f"""
<div style='font-size:.72rem;color:#8b98a5'>
📅 {date.today().strftime('%d %b %Y')}<br>
🧾 <b style='color:#00a3ff'>{n}</b> invoice(s) saved<br>
💸 <b style='color:#ffd700'>{ne}</b> expense(s) tracked<br>
🔒 Session active
</div>""",
            unsafe_allow_html=True,
        )

        if st.button("Logout", use_container_width=True):
            for k in ["authenticated", "username", "role", "display_name"]:
                st.session_state.pop(k, None)
            st.rerun()

    return page


# ════════════════════════════════════════════════════════════════════════════
#  SHARED: clean invoice preview table (used by manual + AI pages)
# ════════════════════════════════════════════════════════════════════════════
def render_invoice_preview(inv: dict):
    is_usd = inv.get("usd_amount") not in [None, 0, 0.0]
    badge = (
        "<span class='badge-export'>EXPORT / LUT — 0% GST</span>"
        if is_usd
        else "<span class='badge-inr'>DOMESTIC — 18% GST</span>"
    )
    st.markdown(badge, unsafe_allow_html=True)

    if is_usd:
        rows = [
            ("Client", inv.get("client_name", "—")),
            ("Description", inv.get("description", "—")),
            ("Date", inv.get("date", "—")),
            ("USD Amount", fmt_usd(inv["usd_amount"])),
            ("Exchange Rate", f"₹{inv.get('exchange_rate', live_rate()):.2f}"),
            ("INR Equivalent", fmt_inr(inv["inr_amount"])),
            ("GST", inv["gst_type"]),
            ("TDS", f"{fmt_inr(inv['tds_amount'])} ({inv['tds_rate']*100:.0f}%)"),
            ("✅ Net Receivable", fmt_inr(inv["net_receivable"])),
        ]
    else:
        rows = [
            ("Client", inv.get("client_name", "—")),
            ("Description", inv.get("description", "—")),
            ("Date", inv.get("date", "—")),
            ("Taxable Value", fmt_inr(inv["inr_amount"])),
            ("CGST (9%)", fmt_inr(inv["cgst"])),
            ("SGST (9%)", fmt_inr(inv["sgst"])),
            ("Total GST", fmt_inr(inv["total_gst"])),
            ("Invoice Total", fmt_inr(inv.get("invoice_total", 0))),
            ("TDS", f"{fmt_inr(inv['tds_amount'])} ({inv['tds_rate']*100:.0f}%)"),
            ("✅ Net Receivable", fmt_inr(inv["net_receivable"])),
        ]
    df = pd.DataFrame(rows, columns=["Field", "Value"])
    st.dataframe(df, use_container_width=True, hide_index=True)


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: DASHBOARD
# ════════════════════════════════════════════════════════════════════════════
def page_dashboard():
    st.markdown("<h2>📊 Financial Dashboard</h2>", unsafe_allow_html=True)
    inv = st.session_state["invoices"]

    total_rev = sum(safe_float(i.get("inr_amount")) for i in inv)
    total_gst = sum(safe_float(i.get("total_gst")) for i in inv)
    total_tds = sum(safe_float(i.get("tds_amount")) for i in inv)
    total_net = sum(safe_float(i.get("net_receivable")) for i in inv)
    usd_inv = [i for i in inv if i.get("usd_amount") not in [None, 0, 0.0]]

    c1, c2, c3, c4 = st.columns(4)
    for col, lbl, val, sub in [
        (c1, "TOTAL REVENUE", fmt_inr(total_rev), f"{len(inv)} invoices"),
        (c2, "GST LIABILITY", fmt_inr(total_gst), "CGST+SGST 18%"),
        (c3, "TDS RECEIVABLE", fmt_inr(total_tds), "Claim in ITR"),
        (c4, "NET RECEIVABLE", fmt_inr(total_net), "After TDS"),
    ]:
        with col:
            st.markdown(
                f"""
<div class='metric-card'>
    <div class='metric-label'>{lbl}</div>
    <div class='metric-value'>{val}</div>
    <div class='metric-sub'>{sub}</div>
</div>""",
                unsafe_allow_html=True,
            )

    st.divider()
    ca, cb = st.columns(2)
    with ca:
        st.markdown(
            "<div class='section-header'>Revenue Breakdown</div>",
            unsafe_allow_html=True,
        )
        if inv:
            df = pd.DataFrame(
                {
                    "Inv": [f"#{i+1}" for i in range(len(inv))],
                    "Taxable": [safe_float(x.get("inr_amount")) for x in inv],
                    "GST": [safe_float(x.get("total_gst")) for x in inv],
                    "TDS": [safe_float(x.get("tds_amount")) for x in inv],
                }
            ).set_index("Inv")
            st.bar_chart(df, height=220)
        else:
            st.info("No invoices yet — create one in 'New Invoice'.")
    with cb:
        st.markdown(
            "<div class='section-header'>Export vs Domestic</div>",
            unsafe_allow_html=True,
        )
        ex = sum(
            safe_float(i.get("inr_amount"))
            for i in inv
            if i.get("usd_amount") not in [None, 0, 0.0]
        )
        dm = sum(
            safe_float(i.get("inr_amount"))
            for i in inv
            if i.get("usd_amount") in [None, 0, 0.0]
        )
        if ex + dm > 0:
            st.bar_chart(
                pd.DataFrame(
                    {"Category": ["Export", "Domestic"], "Amount": [ex, dm]}
                ).set_index("Category"),
                height=220,
            )
        else:
            st.info("Will populate after invoices are added.")

    # Client-wise summary
    if inv:
        st.markdown(
            "<div class='section-header'>👤 Client-Wise Summary</div>",
            unsafe_allow_html=True,
        )
        cl = {}
        for i in inv:
            cn = i.get("client_name", "Unknown")
            if cn not in cl:
                cl[cn] = {
                    "Revenue": 0.0,
                    "GST": 0.0,
                    "TDS": 0.0,
                    "Net": 0.0,
                    "Count": 0,
                }
            cl[cn]["Revenue"] += safe_float(i.get("inr_amount"))
            cl[cn]["GST"] += safe_float(i.get("total_gst"))
            cl[cn]["TDS"] += safe_float(i.get("tds_amount"))
            cl[cn]["Net"] += safe_float(i.get("net_receivable"))
            cl[cn]["Count"] += 1
        rows = [
            {
                "Client": cn,
                "#": d["Count"],
                "Revenue": fmt_inr(d["Revenue"]),
                "GST": fmt_inr(d["GST"]),
                "TDS": fmt_inr(d["TDS"]),
                "Net": fmt_inr(d["Net"]),
            }
            for cn, d in cl.items()
        ]
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    if usd_inv:
        st.markdown(
            "<div class='section-header'>🌐 Export of Service (LUT)</div>",
            unsafe_allow_html=True,
        )
        st.markdown(
            "<span class='badge-export'>0% GST — LUT Filed</span> <span style='font-size:.78rem;color:#8b98a5;margin-left:8px'>IGST Act § 16(1)(a)</span>",
            unsafe_allow_html=True,
        )
        st.dataframe(
            pd.DataFrame(
                [
                    {
                        "Date": i["date"],
                        "Client": i.get("client_name", "—"),
                        "USD": fmt_usd(i["usd_amount"]),
                        "INR": fmt_inr(i["inr_amount"]),
                        "Rate": f"₹{i.get('exchange_rate',live_rate()):.2f}",
                        "Desc": i.get("description", "—"),
                    }
                    for i in usd_inv
                ]
            ),
            use_container_width=True,
            hide_index=True,
        )


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: NEW INVOICE  (manual entry — unchanged core logic, now uses live_rate)
# ════════════════════════════════════════════════════════════════════════════
def page_new_invoice():
    st.markdown("<h2>🧾 Create Invoice</h2>", unsafe_allow_html=True)
    c1, c2 = st.columns([1.2, 1])

    with c1:
        st.markdown(
            "<div class='section-header'>Invoice Details</div>", unsafe_allow_html=True
        )

        # v3: pre-select from sidebar default
        mode_opts = ["USD (Export)", "INR (Domestic)"]
        mode_idx = 0 if st.session_state["default_currency"] == "USD (Export)" else 1
        mode = st.selectbox("Currency Mode", mode_opts, index=mode_idx)
        client_name = st.text_input("Client Name *", placeholder="e.g. Acme Corp")
        description = st.text_input(
            "Work Description", placeholder="e.g. Web Dev – Q1 2026"
        )
        amount = st.number_input(
            f"Amount ({'USD' if 'USD' in mode else 'INR'})",
            min_value=0.0,
            value=1000.0,
            step=100.0,
            format="%.2f",
        )
        tds_label = st.selectbox("TDS Section", list(TDS_RATES.keys()))

        # Show live rate reminder for USD mode
        if "USD" in mode:
            st.markdown(
                f"<div class='rate-chip' style='margin-bottom:.5rem'>Using ₹{live_rate():.2f}/USD (change in sidebar)</div>",
                unsafe_allow_html=True,
            )

        if st.button("⚡  Calculate & Save Invoice", use_container_width=True):
            if not client_name.strip():
                st.warning("⚠️ Client Name is required.")
            else:
                inv = calculate_invoice(
                    mode, amount, tds_label, description, client_name
                )
                st.session_state["current_invoice"] = inv
                save_invoice(inv)
                st.success(
                    f"✅ Invoice saved! {len(st.session_state['invoices'])} total invoices."
                )

    with c2:
        st.markdown(
            "<div class='section-header'>Live Preview</div>", unsafe_allow_html=True
        )
        inv = st.session_state.get("current_invoice")
        if inv:
            render_invoice_preview(inv)
        else:
            st.info("Calculate an invoice to see the preview here.")


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: INVOICE HISTORY
# ════════════════════════════════════════════════════════════════════════════
def page_history():
    st.markdown("<h2>📁 Invoice History</h2>", unsafe_allow_html=True)
    inv = st.session_state["invoices"]
    if not inv:
        st.info("No invoices yet — head to New Invoice.")
        return

    clients = sorted(set(i.get("client_name", "Unknown") for i in inv))
    sel = st.selectbox("Filter by Client", ["All"] + clients)
    filtered = inv if sel == "All" else [i for i in inv if i.get("client_name") == sel]

    rows = [
        {
            "#": idx + 1,
            "Date": i["date"],
            "Client": i.get("client_name", "—"),
            "Desc": i.get("description", "—"),
            "Mode": i["mode"],
            "Taxable(₹)": safe_float(i.get("inr_amount")),
            "GST(₹)": safe_float(i.get("total_gst")),
            "TDS(₹)": safe_float(i.get("tds_amount")),
            "Net(₹)": safe_float(i.get("net_receivable")),
        }
        for idx, i in enumerate(filtered)
    ]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    st.caption(f"Showing {len(filtered)} of {len(inv)} invoices")

    if st.button("🗑  Clear All Invoices"):
        st.session_state["invoices"] = []
        st.session_state["current_invoice"] = None
        st.session_state["ai_extracted"] = None
        st.rerun()


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: AI ASSISTANT  — v3 fully synced with manual logic
# ════════════════════════════════════════════════════════════════════════════
def page_ai():
    st.markdown("<h2>🤖 AI Financial Assistant</h2>", unsafe_allow_html=True)
    st.markdown(
        "<div style='color:#8b98a5;font-size:.82rem'>Gemini 1.5 Flash · Multi-modal PDF/Image · Synced with Manual Entry Logic</div>",
        unsafe_allow_html=True,
    )

    model = get_gemini()
    if not model:
        st.warning("⚠️ GEMINI_API_KEY missing in Streamlit Secrets. Add it to continue.")
        return

    tab1, tab2 = st.tabs(["💬 Chat", "📄 Scan Invoice → Dashboard"])

    # ── Tab 1: Chat ──────────────────────────────────────────────────────────
    with tab1:
        q = st.text_area(
            "Ask about GST, TDS, ITR, invoicing...",
            height=100,
            placeholder="e.g. Section 44ADA kya hai? GSTR-1 monthly ya quarterly?",
        )
        inv = st.session_state["invoices"]
        ctx = ""
        if inv:
            ctx = (
                f"\n\nUser context: {len(inv)} invoices saved. "
                f"Total Revenue ₹{sum(safe_float(i.get('inr_amount')) for i in inv):,.0f}, "
                f"GST ₹{sum(safe_float(i.get('total_gst')) for i in inv):,.0f}. "
                f"USD Rate: ₹{live_rate():.2f}/USD."
            )
        if st.button("Ask Gemini ⚡"):
            if q.strip():
                with st.spinner("Consulting AI..."):
                    prompt = (
                        "You are DataSnap Pro's AI financial advisor for Indian freelancers. "
                        "Give accurate, concise advice on GST/TDS/ITR. Cite relevant sections."
                        + ctx
                        + "\n\nUser: "
                        + q
                    )
                    resp = model.generate_content(prompt)
                    st.markdown(
                        f"<div class='ai-response'>{resp.text}</div>",
                        unsafe_allow_html=True,
                    )

    # ── Tab 2: Invoice Scanner ───────────────────────────────────────────────
    with tab2:
        st.markdown(
            "<div class='section-header'>Scan Invoice → AI Extracts → Confirm → Dashboard</div>",
            unsafe_allow_html=True,
        )

        # Show current rate being used
        st.markdown(
            f"<div class='info-box'>ℹ️ AI will use the live rate <b style='color:#00a3ff'>₹{live_rate():.2f}/USD</b> "
            f"(set in sidebar) for all calculations after extraction.</div>",
            unsafe_allow_html=True,
        )

        uploaded = st.file_uploader(
            "Upload Invoice PDF or Image",
            type=["pdf", "png", "jpg", "jpeg"],
        )

        if uploaded:
            fb = uploaded.read()
            mime = uploaded.type
            st.success(f"✅ {uploaded.name}  ({len(fb)//1024} KB loaded)")

            if st.button("🔍 Extract with Gemini"):
                with st.spinner("Reading invoice with AI..."):
                    b64 = base64.standard_b64encode(fb).decode()

                    # v3: Tell AI which currency to detect, and current rate
                    prompt = f"""You are a financial data extraction engine for Indian freelancers.
Current USD to INR exchange rate: {live_rate():.2f}

Extract data from this invoice and return ONLY valid JSON (no markdown, no backticks, no extra text):

{{
  "client_name": "Company or person name",
  "description": "Work or service description",
  "date": "DD Mon YYYY",
  "detected_currency": "USD" or "INR",
  "mode": "USD (Export)" or "INR (Domestic)",
  "usd_amount": 500.0 or null,
  "inr_amount": 46540.0,
  "cgst": 0.0,
  "sgst": 0.0,
  "total_gst": 0.0,
  "tds_rate": 0.0,
  "tds_amount": 0.0,
  "net_receivable": 46540.0
}}

Rules:
- If invoice has USD amounts → mode is "USD (Export)", set usd_amount, convert inr_amount using rate {live_rate():.2f}
- If invoice has INR amounts → mode is "INR (Domestic)", set usd_amount to null
- If GST is mentioned → split into cgst and sgst (9% each)
- Set all missing numeric fields to 0.0, never null for numbers
- Return ONLY the JSON object, nothing else"""

                    try:
                        resp = model.generate_content(
                            [
                                {"inline_data": {"mime_type": mime, "data": b64}},
                                {"text": prompt},
                            ]
                        )
                        raw_text = resp.text.strip()

                        # Strip any accidental markdown fences
                        for fence in ["```json", "```python", "```"]:
                            raw_text = raw_text.replace(fence, "")
                        raw_text = raw_text.strip()

                        parsed = json.loads(raw_text)

                        # FIX #4 — null-guard ALL numeric fields before any arithmetic
                        numeric_fields = [
                            "usd_amount",
                            "inr_amount",
                            "cgst",
                            "sgst",
                            "total_gst",
                            "tds_rate",
                            "tds_amount",
                            "net_receivable",
                        ]
                        for field in numeric_fields:
                            parsed[field] = safe_float(parsed.get(field))

                        # FIX #1: Run through canonical calculator for consistency
                        normalised = normalise_ai_extraction(parsed)
                        st.session_state["ai_extracted"] = normalised

                    except json.JSONDecodeError as e:
                        st.error(f"⚠️ AI returned non-JSON response. Error: {e}")
                        st.markdown(
                            "<div class='ai-response'>Try uploading a clearer image or PDF with visible text.</div>",
                            unsafe_allow_html=True,
                        )
                        st.session_state["ai_extracted"] = None
                    except Exception as e:
                        st.error(f"⚠️ Extraction failed: {e}")
                        st.session_state["ai_extracted"] = None

        # ── AI BRIDGE CARD — outside upload block so it survives reruns ──────
        extracted = st.session_state.get("ai_extracted")
        if extracted:
            st.markdown("<div class='bridge-card'>", unsafe_allow_html=True)
            st.markdown(
                "<div class='bridge-title'>🤖 AI Extracted Data — Review & Confirm</div>",
                unsafe_allow_html=True,
            )

            # FIX #5 — clean st.dataframe, no raw JSON shown ──────────────────
            is_usd = extracted.get("usd_amount") not in [None, 0, 0.0]

            # Editable fields before confirming
            st.markdown("**✏️ You can adjust these before saving:**")
            col_e1, col_e2 = st.columns(2)
            with col_e1:
                ai_client = st.text_input(
                    "Client Name",
                    value=extracted.get("client_name", ""),
                    key="ai_client_edit",
                )
                ai_desc = st.text_input(
                    "Description",
                    value=extracted.get("description", ""),
                    key="ai_desc_edit",
                )
                ai_mode_opts = ["USD (Export)", "INR (Domestic)"]
                ai_mode_idx = 0 if is_usd else 1
                ai_mode = st.selectbox(
                    "Currency Mode",
                    ai_mode_opts,
                    index=ai_mode_idx,
                    key="ai_mode_edit",
                )
            with col_e2:
                ai_amount = st.number_input(
                    f"Amount ({'USD' if 'USD' in ai_mode else 'INR'})",
                    min_value=0.0,
                    value=float(
                        extracted.get("usd_amount", 0) or extracted.get("inr_amount", 0)
                    ),
                    step=100.0,
                    format="%.2f",
                    key="ai_amount_edit",
                )
                ai_tds = st.selectbox(
                    "TDS Section",
                    list(TDS_RATES.keys()),
                    key="ai_tds_edit",
                )
                ai_date = st.text_input(
                    "Invoice Date",
                    value=extracted.get("date", datetime.now().strftime("%d %b %Y")),
                    key="ai_date_edit",
                )

            # Recalculate preview using edited values + live rate
            preview_inv = calculate_invoice(
                mode=ai_mode,
                amount=ai_amount,
                tds_label=ai_tds,
                description=ai_desc,
                client_name=ai_client,
            )
            preview_inv["date"] = ai_date

            st.markdown(
                "<br>**📋 Calculated Preview (using live rate):**",
                unsafe_allow_html=True,
            )
            render_invoice_preview(preview_inv)  # FIX #5: clean table, no raw JSON

            st.markdown("<br>", unsafe_allow_html=True)
            col_confirm, col_discard = st.columns(2)

            with col_confirm:
                # FIX #1: same save_invoice() as manual page
                if st.button(
                    "✅ Confirm & Add to Dashboard",
                    use_container_width=True,
                    type="primary",
                ):
                    save_invoice(preview_inv)
                    st.session_state["current_invoice"] = preview_inv
                    st.session_state["ai_extracted"] = None
                    st.success(
                        f"🚀 Invoice added! Total: {len(st.session_state['invoices'])} invoices in Dashboard."
                    )
                    st.rerun()

            with col_discard:
                if st.button("🗑 Discard", use_container_width=True):
                    st.session_state["ai_extracted"] = None
                    st.rerun()

            st.markdown("</div>", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: EXPORT & WHATSAPP
# ════════════════════════════════════════════════════════════════════════════
def page_export():
    st.markdown("<h2>📤 Export & Notifications</h2>", unsafe_allow_html=True)
    inv = st.session_state["invoices"]
    tab1, tab2 = st.tabs(["📊 Excel Report", "💬 WhatsApp"])

    with tab1:
        st.markdown(
            "<div class='section-header'>CA Audit Excel Report</div>",
            unsafe_allow_html=True,
        )
        if not inv:
            st.info("No invoices yet — create some first.")
        else:
            st.success(f"✅ {len(inv)} invoice(s) ready to export.")
            owner = st.text_input(
                "Freelancer / Owner Name",
                value=st.session_state.get("display_name", ""),
            )
            xls = build_excel(inv, owner or "—")
            st.download_button(
                label=f"📥 Download Excel — {len(inv)} invoices",
                data=xls,
                file_name=f"DataSnap_Audit_{date.today().isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.caption("3 sheets: CA Audit · Client Summary · Overall Summary")

    with tab2:
        st.markdown(
            "<div class='section-header'>WhatsApp Notification</div>",
            unsafe_allow_html=True,
        )
        st.markdown(
            """<div class='info-box'>
Pre-wired for <b style='color:#00a3ff'>Twilio / WATI</b> API.
Add <code>TWILIO_SID</code>, <code>TWILIO_TOKEN</code> or <code>WATI_TOKEN</code> to Secrets and uncomment code in <code>send_whatsapp_placeholder()</code>.
</div>""",
            unsafe_allow_html=True,
        )
        phone = st.text_input("Client WhatsApp Number", placeholder="+91XXXXXXXXXX")
        c_name = st.text_input("Client Name for Message", value="Client")
        sel_inv = st.session_state.get("current_invoice") or (inv[-1] if inv else None)
        if sel_inv:
            msg = whatsapp_message(sel_inv, c_name, phone)
            st.markdown("**Preview:**")
            st.markdown(
                f"<div class='whatsapp-preview'>{msg}</div>", unsafe_allow_html=True
            )
            if st.button("📲 Send (Mock)"):
                r = send_whatsapp_placeholder(phone, msg)
                st.success(f"Mock → {r['to']} | {r['chars']} chars | {r['status']}")
                st.info(
                    "To go live: uncomment Twilio/WATI block in send_whatsapp_placeholder()."
                )
        else:
            st.info("Create an invoice first to generate a WhatsApp preview.")


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: ADMIN
# ════════════════════════════════════════════════════════════════════════════
def page_admin():
    st.markdown("<h2>👥 Admin Panel</h2>", unsafe_allow_html=True)
    st.markdown(
        "<div class='section-header'>Registered Users</div>", unsafe_allow_html=True
    )
    st.dataframe(
        pd.DataFrame(
            [
                {"Username": u, "Name": d["name"], "Role": d["role"]}
                for u, d in USERS.items()
            ]
        ),
        use_container_width=True,
        hide_index=True,
    )
    st.info(
        "To add users: edit the USERS dict or wire to SQLite / Supabase for persistence."
    )
    st.markdown(
        "<div class='section-header'>Live Configuration</div>", unsafe_allow_html=True
    )
    st.json(
        {
            "Live_USD_Rate": f"₹{live_rate():.2f}",
            "Default_Currency": st.session_state["default_currency"],
            "CGST": "9%",
            "SGST": "9%",
            "GST_Total": "18%",
            "AI_Model": "gemini-1.5-flash",
            "GSheets_Sync": "stub (uncomment to activate)",
            "Invoices_This_Session": len(st.session_state["invoices"]),
        }
    )


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: EXPENSE TRACKER  (v4 new)
# ════════════════════════════════════════════════════════════════════════════
def page_expenses():
    st.markdown("<h2>💸 Expense Tracker</h2>", unsafe_allow_html=True)
    st.markdown(
        "<div style='color:#8b98a5;font-size:.82rem'>Track business expenses · See Profit/Loss · Export to CA</div>",
        unsafe_allow_html=True,
    )

    tab_add, tab_view, tab_pl = st.tabs(
        ["➕ Add Expense", "📋 All Expenses", "📈 Profit & Loss"]
    )

    # ── Tab 1: Add Expense ────────────────────────────────────────────────────
    with tab_add:
        st.markdown(
            "<div class='section-header'>New Expense Entry</div>",
            unsafe_allow_html=True,
        )
        c1, c2 = st.columns(2)
        with c1:
            exp_date = st.date_input("Date", value=date.today(), key="exp_date")
            exp_cat = st.selectbox("Category", EXPENSE_CATEGORIES, key="exp_cat")
            exp_vendor = st.text_input(
                "Vendor / Paid To",
                placeholder="e.g. Adobe, AWS, WeWork",
                key="exp_vendor",
            )
        with c2:
            exp_amount = st.number_input(
                "Amount (₹)",
                min_value=0.0,
                value=0.0,
                step=100.0,
                format="%.2f",
                key="exp_amount",
            )
            exp_gst_paid = st.number_input(
                "GST Paid on Expense (₹)",
                min_value=0.0,
                value=0.0,
                step=10.0,
                format="%.2f",
                key="exp_gst_paid",
                help="GST you paid — can be used for Input Tax Credit (ITC)",
            )
            exp_desc = st.text_input(
                "Description / Notes",
                placeholder="e.g. Annual subscription",
                key="exp_desc",
            )

        exp_receipt = st.file_uploader(
            "Receipt (optional)", type=["pdf", "png", "jpg", "jpeg"], key="exp_receipt"
        )

        if st.button("💾 Save Expense", use_container_width=True, type="primary"):
            if exp_amount <= 0:
                st.warning("⚠️ Amount must be greater than 0.")
            else:
                entry = {
                    "date": exp_date.strftime("%d %b %Y"),
                    "category": exp_cat,
                    "vendor": exp_vendor.strip() or "—",
                    "description": exp_desc.strip() or "—",
                    "amount": exp_amount,
                    "gst_paid": exp_gst_paid,
                    "net_expense": exp_amount - exp_gst_paid,
                    "has_receipt": exp_receipt is not None,
                    "saved_by": st.session_state.get("username", ""),
                }
                st.session_state["expenses"].append(entry)
                st.success(
                    f"✅ Expense saved! Total {len(st.session_state['expenses'])} expenses tracked."
                )
                st.rerun()

    # ── Tab 2: View All Expenses ──────────────────────────────────────────────
    with tab_view:
        exps = st.session_state["expenses"]
        if not exps:
            st.info("No expenses yet — add one above.")
        else:
            # Category filter
            cats = sorted(set(e["category"] for e in exps))
            sel_cat = st.selectbox(
                "Filter by Category", ["All"] + cats, key="exp_filter"
            )
            filtered_exps = (
                exps
                if sel_cat == "All"
                else [e for e in exps if e["category"] == sel_cat]
            )

            rows = [
                {
                    "#": i + 1,
                    "Date": e["date"],
                    "Category": e["category"],
                    "Vendor": e["vendor"],
                    "Desc": e["description"],
                    "Amount(₹)": e["amount"],
                    "GST Paid(₹)": e["gst_paid"],
                    "Net(₹)": e["net_expense"],
                    "Receipt": "✅" if e["has_receipt"] else "—",
                }
                for i, e in enumerate(filtered_exps)
            ]
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

            # Category-wise summary
            st.markdown(
                "<div class='section-header'>Category Breakdown</div>",
                unsafe_allow_html=True,
            )
            cat_data = {}
            for e in exps:
                cat_data.setdefault(e["category"], 0)
                cat_data[e["category"]] += e["amount"]
            cat_df = pd.DataFrame(
                {"Category": list(cat_data.keys()), "Amount": list(cat_data.values())}
            ).set_index("Category")
            st.bar_chart(cat_df, height=200)

            total_exp = sum(e["amount"] for e in exps)
            total_itc = sum(e["gst_paid"] for e in exps)
            st.markdown(
                f"""
<div style='display:flex;gap:1rem;margin-top:1rem'>
    <div class='metric-card' style='flex:1'>
        <div class='metric-label'>Total Expenses</div>
        <div class='metric-value'>{fmt_inr(total_exp)}</div>
        <div class='metric-sub'>{len(exps)} entries</div>
    </div>
    <div class='metric-card' style='flex:1;border-top-color:var(--teal)'>
        <div class='metric-label'>ITC Claimable (GST)</div>
        <div class='metric-value' style='color:var(--teal)'>{fmt_inr(total_itc)}</div>
        <div class='metric-sub'>Input Tax Credit</div>
    </div>
</div>""",
                unsafe_allow_html=True,
            )

            if st.button("🗑 Clear All Expenses"):
                st.session_state["expenses"] = []
                st.rerun()

    # ── Tab 3: Profit & Loss ──────────────────────────────────────────────────
    with tab_pl:
        inv = st.session_state["invoices"]
        exps = st.session_state["expenses"]

        total_revenue = sum(safe_float(i.get("inr_amount")) for i in inv)
        total_gst_collected = sum(safe_float(i.get("total_gst")) for i in inv)
        total_tds = sum(safe_float(i.get("tds_amount")) for i in inv)
        total_net_income = sum(safe_float(i.get("net_receivable")) for i in inv)

        total_expenses = sum(e["amount"] for e in exps)
        total_itc = sum(e["gst_paid"] for e in exps)
        gst_payable = max(0, total_gst_collected - total_itc)

        gross_profit = total_revenue - total_expenses
        net_profit = total_net_income - (total_expenses - total_itc)
        profit_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0

        card_cls = "profit-card" if net_profit >= 0 else "loss-card"
        status_color = "var(--green)" if net_profit >= 0 else "var(--red)"
        status_label = "PROFIT" if net_profit >= 0 else "LOSS"

        st.markdown(
            f"""
<div class='{card_cls}'>
    <div class='metric-label'>{status_label} THIS PERIOD</div>
    <div class='metric-value' style='color:{status_color};font-size:2.5rem'>{fmt_inr(abs(net_profit))}</div>
    <div class='metric-sub'>Margin: {profit_margin:.1f}% | After TDS & ITC</div>
</div>""",
            unsafe_allow_html=True,
        )

        c1, c2, c3 = st.columns(3)
        for col, lbl, val, col_override in [
            (c1, "GROSS REVENUE", fmt_inr(total_revenue), None),
            (c2, "TOTAL EXPENSES", fmt_inr(total_expenses), "var(--red)"),
            (c3, "GST PAYABLE (NET)", fmt_inr(gst_payable), "var(--gold)"),
        ]:
            with col:
                st.markdown(
                    f"""
<div class='metric-card'>
    <div class='metric-label'>{lbl}</div>
    <div class='metric-value' style='{"color:" + col_override if col_override else ""}'>{val}</div>
</div>""",
                    unsafe_allow_html=True,
                )

        # P&L Table
        st.markdown(
            "<div class='section-header'>Detailed P&L Statement</div>",
            unsafe_allow_html=True,
        )
        pl_rows = [
            ("📥 Gross Revenue (Taxable)", fmt_inr(total_revenue), "Income"),
            ("📤 GST Collected (Liability)", fmt_inr(total_gst_collected), "Tax"),
            ("🔻 Total Expenses", fmt_inr(total_expenses), "Deduction"),
            ("♻️ ITC (GST on Expenses)", fmt_inr(total_itc), "Credit"),
            ("🏦 Net GST Payable", fmt_inr(gst_payable), "Tax Due"),
            ("🔻 TDS Deducted", fmt_inr(total_tds), "Deduction"),
            ("✅ Net Profit / Loss", fmt_inr(net_profit), status_label),
        ]
        st.dataframe(
            pd.DataFrame(pl_rows, columns=["Line Item", "Amount", "Type"]),
            use_container_width=True,
            hide_index=True,
        )
        st.caption(
            "💡 Share this P&L with your CA for ITR filing. Export full data from Export page."
        )


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: RAG FINANCIAL ASSISTANT  (v4 new)
#  Retrieval-Augmented Generation — answers questions FROM your actual data
# ════════════════════════════════════════════════════════════════════════════
def page_rag():
    st.markdown("<h2>🧠 Smart Data Assistant</h2>", unsafe_allow_html=True)
    st.markdown(
        "<div style='color:#8b98a5;font-size:.82rem'>Ask questions about YOUR invoices & expenses · AI retrieves from your actual data · RAG-powered</div>",
        unsafe_allow_html=True,
    )

    model = get_gemini()
    if not model:
        st.warning("⚠️ GEMINI_API_KEY missing. Add it to Streamlit Secrets.")
        return

    inv = st.session_state["invoices"]
    exps = st.session_state["expenses"]

    if not inv and not exps:
        st.info(
            "💡 Add some invoices and expenses first — then ask me anything about your financial data!"
        )
        return

    # ── Build RAG Context (the "Retrieval" part) ──────────────────────────────
    def build_rag_context() -> str:
        """Serialize all user data into a structured context string for the AI."""
        lines = ["=== FREELANCER FINANCIAL DATA ===\n"]

        # Invoice summary
        lines.append(f"INVOICES ({len(inv)} total):")
        total_rev = sum(safe_float(i.get("inr_amount")) for i in inv)
        total_gst = sum(safe_float(i.get("total_gst")) for i in inv)
        total_tds = sum(safe_float(i.get("tds_amount")) for i in inv)
        total_net = sum(safe_float(i.get("net_receivable")) for i in inv)
        lines.append(f"  - Total Revenue: ₹{total_rev:,.2f}")
        lines.append(f"  - Total GST Collected: ₹{total_gst:,.2f}")
        lines.append(f"  - Total TDS Deducted: ₹{total_tds:,.2f}")
        lines.append(f"  - Total Net Receivable: ₹{total_net:,.2f}")

        # Client breakdown
        clients = {}
        for i in inv:
            cn = i.get("client_name", "Unknown")
            clients.setdefault(cn, {"revenue": 0, "count": 0, "net": 0})
            clients[cn]["revenue"] += safe_float(i.get("inr_amount"))
            clients[cn]["count"] += 1
            clients[cn]["net"] += safe_float(i.get("net_receivable"))
        lines.append("\nCLIENT BREAKDOWN:")
        for cn, d in sorted(clients.items(), key=lambda x: -x[1]["revenue"]):
            lines.append(
                f"  - {cn}: {d['count']} invoice(s), Revenue ₹{d['revenue']:,.2f}, Net ₹{d['net']:,.2f}"
            )

        # Individual invoices
        lines.append("\nINDIVIDUAL INVOICES:")
        for idx, i in enumerate(inv, 1):
            lines.append(
                f"  #{idx} [{i['date']}] {i.get('client_name','?')} | "
                f"{i.get('description','—')} | "
                f"Mode: {i['mode']} | "
                f"Taxable: ₹{safe_float(i.get('inr_amount')):,.2f} | "
                f"Net: ₹{safe_float(i.get('net_receivable')):,.2f}"
            )
            if i.get("usd_amount"):
                lines[-1] += f" | USD: ${safe_float(i.get('usd_amount')):,.2f}"

        # Expense data
        if exps:
            total_exp = sum(e["amount"] for e in exps)
            total_itc = sum(e["gst_paid"] for e in exps)
            net_profit = total_net - (total_exp - total_itc)
            lines.append(f"\nEXPENSES ({len(exps)} total):")
            lines.append(f"  - Total Expenses: ₹{total_exp:,.2f}")
            lines.append(f"  - ITC (GST on expenses): ₹{total_itc:,.2f}")
            lines.append(f"  - Net Profit/Loss: ₹{net_profit:,.2f}")
            lines.append("\nEXPENSE DETAILS:")
            for e in exps:
                lines.append(
                    f"  [{e['date']}] {e['category']} | {e['vendor']} | ₹{e['amount']:,.2f}"
                )

        lines.append(f"\nCurrent USD Rate: ₹{live_rate():.2f}/USD")
        lines.append(f"Report Generated: {datetime.now().strftime('%d %b %Y %H:%M')}")
        return "\n".join(lines)

    # ── Suggested Questions ───────────────────────────────────────────────────
    st.markdown(
        "<div class='section-header'>💡 Try asking...</div>", unsafe_allow_html=True
    )
    suggestions = [
        "Who is my best client by revenue?",
        "What is my total profit this period?",
        "How much GST do I owe after ITC?",
        "Which expense category am I spending most on?",
        "How much TDS can I claim in ITR?",
        "Give me a financial summary for my CA",
        "Which month had the highest income?",
    ]
    cols = st.columns(3)
    for i, sug in enumerate(suggestions[:6]):
        with cols[i % 3]:
            if st.button(sug, use_container_width=True, key=f"sug_{i}"):
                st.session_state["rag_pending_q"] = sug
                st.rerun()

    st.divider()

    # ── Chat Interface ────────────────────────────────────────────────────────
    chat_history = st.session_state["rag_chat_history"]

    # Show history
    for msg in chat_history[-10:]:  # show last 10 exchanges
        if msg["role"] == "user":
            st.markdown(
                f"<div class='rag-msg-user'>🧑 {msg['content']}</div>",
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f"<div class='rag-msg-ai'>🧠 {msg['content']}"
                f"<div class='rag-source'>📊 Answer derived from your {len(inv)} invoices"
                + (f" & {len(exps)} expenses" if exps else "")
                + "</div></div>",
                unsafe_allow_html=True,
            )

    # Input
    pending = st.session_state.pop("rag_pending_q", "")
    user_q = st.text_input(
        "Ask anything about your financial data...",
        value=pending,
        placeholder="e.g. Who paid me the most this year?",
        key="rag_input",
    )

    col_ask, col_clear = st.columns([3, 1])
    with col_ask:
        ask_clicked = st.button(
            "🧠 Ask Smart Assistant ⚡", use_container_width=True, type="primary"
        )
    with col_clear:
        if st.button("🗑 Clear Chat", use_container_width=True):
            st.session_state["rag_chat_history"] = []
            st.rerun()

    if ask_clicked and user_q.strip():
        rag_context = build_rag_context()
        system_prompt = f"""You are DataSnap Pro's Smart Financial Assistant for Indian freelancers.
You ONLY answer questions based on the actual financial data provided below.
Do NOT make up numbers. If data is insufficient, say so clearly.
Be concise, specific, and cite actual numbers from the data.
Format currency as ₹X,XX,XXX.XX

=== USER'S ACTUAL FINANCIAL DATA (retrieved from database) ===
{rag_context}
=== END OF DATA ===

Answer the user's question using ONLY the above data."""

        with st.spinner("🧠 Analyzing your data..."):
            try:
                full_prompt = system_prompt + f"\n\nUser Question: {user_q}"
                resp = model.generate_content(full_prompt)
                answer = resp.text.strip()

                st.session_state["rag_chat_history"].append(
                    {"role": "user", "content": user_q}
                )
                st.session_state["rag_chat_history"].append(
                    {"role": "assistant", "content": answer}
                )
                st.rerun()
            except Exception as e:
                st.error(f"⚠️ AI error: {e}")


# ════════════════════════════════════════════════════════════════════════════
#  PAGE: TAX PLANNER  (v4 new)
# ════════════════════════════════════════════════════════════════════════════
def page_tax():
    st.markdown("<h2>🧮 Tax Planner</h2>", unsafe_allow_html=True)
    st.markdown(
        "<div style='color:#8b98a5;font-size:.82rem'>Advance Tax · Section 44ADA · ITR Deadlines · Annual Tax Estimate</div>",
        unsafe_allow_html=True,
    )

    inv = st.session_state["invoices"]
    exps = st.session_state["expenses"]

    total_revenue = sum(safe_float(i.get("inr_amount")) for i in inv)
    total_expenses = sum(e["amount"] for e in exps)
    total_tds = sum(safe_float(i.get("tds_amount")) for i in inv)
    total_itc = sum(e["gst_paid"] for e in exps)

    tab_adv, tab_44ada, tab_gst, tab_deadlines = st.tabs(
        ["📅 Advance Tax", "📊 Sec 44ADA", "🏦 GST Summary", "⏰ Deadlines"]
    )

    # ── Tab 1: Advance Tax Calculator ────────────────────────────────────────
    with tab_adv:
        st.markdown(
            "<div class='section-header'>Advance Tax Estimator</div>",
            unsafe_allow_html=True,
        )
        st.markdown(
            "<div class='info-box'>Advance tax is due if your total tax liability exceeds ₹10,000/year. Pay in 4 instalments to avoid interest under Sec 234B/234C.</div>",
            unsafe_allow_html=True,
        )

        col1, col2 = st.columns(2)
        with col1:
            projected_annual = st.number_input(
                "Projected Annual Income (₹)",
                min_value=0.0,
                value=max(total_revenue * 2, 0.0),  # rough 2x projection
                step=10000.0,
                format="%.0f",
                help="Estimate your full-year taxable income",
            )
            regime = st.radio(
                "Tax Regime", ["New Regime (Default)", "Old Regime"], horizontal=True
            )
        with col2:
            deductions_80c = st.number_input(
                "80C Deductions (₹) — Old Regime only",
                min_value=0.0,
                max_value=150000.0,
                value=0.0,
                step=1000.0,
                disabled="New" in regime,
            )
            hra_exempt = st.number_input(
                "HRA / Other Exemptions (₹)",
                min_value=0.0,
                value=0.0,
                step=1000.0,
                disabled="New" in regime,
            )

        # Calculate tax
        taxable = projected_annual
        if "Old" in regime:
            std_ded = 50000
            taxable = max(0, projected_annual - std_ded - deductions_80c - hra_exempt)

        def calc_tax_new(income):
            slabs = [
                (300000, 0),
                (400000, 0.05),
                (300000, 0.10),
                (300000, 0.15),
                (300000, 0.20),
                (300000, 0.25),
                (float("inf"), 0.30),
            ]
            tax, remaining = 0, income
            for slab_limit, rate in slabs:
                if remaining <= 0:
                    break
                taxable_in_slab = min(remaining, slab_limit)
                tax += taxable_in_slab * rate
                remaining -= taxable_in_slab
            return tax

        def calc_tax_old(income):
            if income <= 250000:
                return 0
            elif income <= 500000:
                return (income - 250000) * 0.05
            elif income <= 1000000:
                return 12500 + (income - 500000) * 0.20
            else:
                return 112500 + (income - 1000000) * 0.30

        if "New" in regime:
            base_tax = calc_tax_new(taxable)
            rebate = min(base_tax, 25000) if taxable <= 700000 else 0
            base_tax = max(0, base_tax - rebate)
        else:
            base_tax = calc_tax_old(taxable)

        surcharge = base_tax * 0.10 if projected_annual > 5000000 else 0
        health_ed_cess = (base_tax + surcharge) * 0.04
        gross_tax = base_tax + surcharge + health_ed_cess
        tax_after_tds = max(0, gross_tax - total_tds)

        # Display
        lbl = "ESTIMATED TAX PAYABLE" if tax_after_tds > 0 else "NO TAX DUE"
        card_cls = "loss-card" if tax_after_tds > 0 else "profit-card"
        color = "var(--red)" if tax_after_tds > 0 else "var(--green)"
        st.markdown(
            f"""
<div class='{card_cls}' style='margin-top:1rem'>
    <div class='metric-label'>{lbl} (After TDS credit)</div>
    <div class='metric-value' style='color:{color}'>{fmt_inr(tax_after_tds)}</div>
    <div class='metric-sub'>Gross Tax: {fmt_inr(gross_tax)} | TDS Credit: {fmt_inr(total_tds)}</div>
</div>""",
            unsafe_allow_html=True,
        )

        if tax_after_tds > 10000:
            st.markdown(
                "<div class='section-header'>📅 Advance Tax Schedule</div>",
                unsafe_allow_html=True,
            )
            instalments = [
                ("15 Jun", 0.15, "1st instalment"),
                ("15 Sep", 0.45, "Cumulative 45%"),
                ("15 Dec", 0.75, "Cumulative 75%"),
                ("15 Mar", 1.00, "100% — Final"),
            ]
            rows = [
                {
                    "Due Date": f"{d} (FY 2025-26)",
                    "Cumulative %": pct,
                    "Amount Due (₹)": fmt_inr(tax_after_tds * pct),
                    "Notes": note,
                }
                for d, pct, note in instalments
            ]
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        else:
            st.success(
                "✅ Advance tax not required (liability below ₹10,000). Pay as Self-Assessment Tax before ITR filing."
            )

    # ── Tab 2: Section 44ADA ──────────────────────────────────────────────────
    with tab_44ada:
        st.markdown(
            "<div class='section-header'>Section 44ADA — Presumptive Taxation</div>",
            unsafe_allow_html=True,
        )
        st.markdown(
            """<div class='info-box'>
<b style='color:#00a3ff'>Who can use 44ADA?</b> Professionals (IT, Design, Consulting, Legal, Medical, etc.) with gross receipts ≤ ₹75 lakh/year.<br>
<b style='color:#00e676'>Benefit:</b> Declare 50% of gross receipts as profit — no need to maintain books of accounts or prove expenses.
</div>""",
            unsafe_allow_html=True,
        )
        ann_rev_44ada = st.number_input(
            "Gross Annual Receipts (₹)",
            min_value=0.0,
            value=max(total_revenue, 0.0),
            step=10000.0,
            format="%.0f",
            key="ada_rev",
        )
        if ann_rev_44ada <= 7500000:
            presumptive_profit = ann_rev_44ada * 0.50
            st.markdown(
                f"""
<div class='profit-card'>
    <div class='metric-label'>PRESUMPTIVE PROFIT (50% of Receipts)</div>
    <div class='metric-value' style='color:var(--green)'>{fmt_inr(presumptive_profit)}</div>
    <div class='metric-sub'>This is your taxable income under 44ADA — file ITR-4</div>
</div>""",
                unsafe_allow_html=True,
            )
            st.markdown(
                f"""
<div class='info-box' style='margin-top:.5rem'>
✅ <b>Your gross receipts ({fmt_inr(ann_rev_44ada)}) are within the ₹75L limit</b> — you qualify for 44ADA.<br>
📝 File <b style='color:#00a3ff'>ITR-4 (Sugam)</b> · No audit required · No expense proof needed<br>
💡 Actual expenses tracked in DataSnap: {fmt_inr(total_expenses)} — compare to see if 44ADA saves you tax.
</div>""",
                unsafe_allow_html=True,
            )
        else:
            st.warning(
                f"⚠️ Gross receipts ({fmt_inr(ann_rev_44ada)}) exceed ₹75L limit. 44ADA not applicable — file ITR-3 with audit."
            )

    # ── Tab 3: GST Summary ────────────────────────────────────────────────────
    with tab_gst:
        st.markdown(
            "<div class='section-header'>GST Liability Summary</div>",
            unsafe_allow_html=True,
        )
        total_gst = sum(safe_float(i.get("total_gst")) for i in inv)
        net_gst = max(0, total_gst - total_itc)

        c1, c2, c3 = st.columns(3)
        for col, lbl, val, color in [
            (c1, "GST Collected", fmt_inr(total_gst), None),
            (c2, "ITC (Input Credit)", fmt_inr(total_itc), "var(--teal)"),
            (c3, "Net GST Payable", fmt_inr(net_gst), "var(--gold)"),
        ]:
            with col:
                st.markdown(
                    f"""
<div class='metric-card'>
    <div class='metric-label'>{lbl}</div>
    <div class='metric-value' style='{"color:" + color if color else ""}'>{val}</div>
</div>""",
                    unsafe_allow_html=True,
                )

        st.markdown(
            "<div class='section-header'>GSTR Filing Guide</div>",
            unsafe_allow_html=True,
        )
        gstr_rows = [
            (
                "GSTR-1",
                "Monthly/Quarterly",
                "Outward supplies (invoices raised)",
                "11th of next month / Quarterly",
            ),
            (
                "GSTR-3B",
                "Monthly",
                "Self-assessed GST return + payment",
                "20th of next month",
            ),
            ("GSTR-9", "Annual", "Annual GST return", "31 December"),
        ]
        st.dataframe(
            pd.DataFrame(
                gstr_rows, columns=["Form", "Frequency", "Purpose", "Due Date"]
            ),
            use_container_width=True,
            hide_index=True,
        )
        st.info(
            "💡 Export revenue = LUT filed → 0% GST. Still file GSTR-1 reporting these invoices under 'Zero-rated supply'."
        )

    # ── Tab 4: Important Deadlines ────────────────────────────────────────────
    with tab_deadlines:
        st.markdown(
            "<div class='section-header'>Key Tax Deadlines FY 2025-26</div>",
            unsafe_allow_html=True,
        )
        deadlines = [
            ("15 Jun 2025", "Advance Tax 1st Instalment (15%)", "⚠️ Upcoming"),
            ("31 Jul 2025", "ITR Filing (Non-audit cases)", "📋 ITR Due"),
            ("15 Sep 2025", "Advance Tax 2nd Instalment (45%)", "⚠️ Upcoming"),
            ("30 Sep 2025", "Tax Audit Completion", "🔍 Audit"),
            ("31 Oct 2025", "ITR Filing (Audit cases)", "📋 ITR Due"),
            ("15 Dec 2025", "Advance Tax 3rd Instalment (75%)", "⚠️ Upcoming"),
            ("15 Mar 2026", "Advance Tax Final Instalment (100%)", "⚠️ Final"),
            ("31 Mar 2026", "Last date to file updated ITR (ITR-U)", "🔄 ITR-U"),
            ("30 Jun 2026", "GSTR-9 Annual Return FY 2024-25", "🏦 GST"),
        ]
        st.dataframe(
            pd.DataFrame(deadlines, columns=["Date", "Event", "Type"]),
            use_container_width=True,
            hide_index=True,
        )
        st.markdown(
            "<div class='tax-pill'>💡 Set calendar reminders for all advance tax dates to avoid 1% monthly interest under Sec 234C</div>",
            unsafe_allow_html=True,
        )


# ════════════════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════════════════
def main():
    inject_css()
    # init_state() already ran at module top

    if not st.session_state["authenticated"]:
        login_block()
        return

    page = render_sidebar()

    if "Dashboard" in page:
        page_dashboard()
    elif "Invoice" in page:
        page_new_invoice()
    elif "History" in page:
        page_history()
    elif "Smart" in page:
        page_rag()
    elif "AI" in page:
        page_ai()
    elif "Expense" in page:
        page_expenses()
    elif "Tax" in page:
        page_tax()
    elif "Export" in page:
        page_export()
    elif "Admin" in page:
        page_admin()


if __name__ == "__main__":
    main()
